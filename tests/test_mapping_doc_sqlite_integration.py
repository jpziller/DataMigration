"""Integration coverage for mapping_doc.py's generate_mapping_workbook()
against a real SQLite file -- previously untested at all (found in
review): every other test in test_mapping_doc.py exercises pure/offline
helpers (extract_insert_columns, check_mapping_balance, set_transform_script)
against a hand-built workbook, never generate_mapping_workbook()'s own
engine-reading path (source_table column introspection via
sql_dialect.list_columns(), and the FieldProfile pre-fill join).
"""
import openpyxl
import pandas as pd
import pytest

import mapping_doc as md
import sql_client
from config import Settings


@pytest.fixture
def sqlite_engine(tmp_path):
    s = Settings(
        sql_backend="sqlite",
        sql_sqlite_dir=str(tmp_path / "_sqlite"),
        sql_sqlite_schemas="dbo",
    )
    return sql_client.make_engine(s), s


def test_generate_mapping_workbook_lists_real_source_columns(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    pd.DataFrame({"Name": ["Acme"], "Legacy_Id": ["A1"], "City": ["Springfield"]}).to_sql(
        "SourceAccounts", engine, schema="dbo", if_exists="replace", index=False
    )
    output_path = str(tmp_path / "mapping.xlsx")

    md.generate_mapping_workbook(None, "Account", output_path, engine, "SourceAccounts", schema="dbo")

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Account"]
    assert ws.cell(row=1, column=2).value == "SourceAccounts"
    assert ws.cell(row=1, column=4).value == "Account"
    source_fields = [ws.cell(row=r, column=2).value for r in range(4, 4 + 3)]
    assert set(source_fields) == {"Name", "Legacy_Id", "City"}


def test_generate_mapping_workbook_raises_on_missing_source_table(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    output_path = str(tmp_path / "mapping.xlsx")
    with pytest.raises(ValueError, match="No such table"):
        md.generate_mapping_workbook(None, "Account", output_path, engine, "DoesNotExist", schema="dbo")


def test_generate_mapping_workbook_prefills_from_existing_field_profile(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    pd.DataFrame({"Name": ["Acme", "Globex"], "Legacy_Id": ["A1", None]}).to_sql(
        "SourceAccounts", engine, schema="dbo", if_exists="replace", index=False
    )
    pd.DataFrame([
        {"ObjectOrTable": "SourceAccounts", "SourceType": "sql_table", "FieldName": "Name",
         "TotalRows": 2, "PopulatedCount": 2, "PopulatedPct": 100.0},
        {"ObjectOrTable": "SourceAccounts", "SourceType": "sql_table", "FieldName": "Legacy_Id",
         "TotalRows": 2, "PopulatedCount": 1, "PopulatedPct": 50.0},
    ]).to_sql("FieldProfile", engine, schema="dbo", if_exists="replace", index=False)

    output_path = str(tmp_path / "mapping.xlsx")
    md.generate_mapping_workbook(None, "Account", output_path, engine, "SourceAccounts", schema="dbo")

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Account"]
    rows_by_field = {
        ws.cell(row=r, column=2).value: r
        for r in range(4, 6)
    }
    name_row = rows_by_field["Name"]
    assert ws.cell(row=name_row, column=6).value == "2 of 2"
    assert ws.cell(row=name_row, column=7).value == 100.0
    legacy_row = rows_by_field["Legacy_Id"]
    assert ws.cell(row=legacy_row, column=6).value == "1 of 2"
    assert ws.cell(row=legacy_row, column=7).value == 50.0


def test_generate_mapping_workbook_regenerates_sheet_in_place_on_second_call(sqlite_engine, tmp_path):
    """Unlike migration_run_book.py's refuse-to-overwrite convention,
    mapping_doc.py regenerates a sheet in place (module docstring/
    check-mapping-balance workflow both assume re-running generate-mapping-doc
    is safe) -- a second call with a changed source table must replace,
    not duplicate, the sheet."""
    engine, _ = sqlite_engine
    pd.DataFrame({"Name": ["Acme"]}).to_sql("SourceAccounts", engine, schema="dbo", if_exists="replace", index=False)
    output_path = str(tmp_path / "mapping.xlsx")
    md.generate_mapping_workbook(None, "Account", output_path, engine, "SourceAccounts", schema="dbo")

    pd.DataFrame({"Name": ["Acme"], "City": ["Springfield"]}).to_sql(
        "SourceAccounts", engine, schema="dbo", if_exists="replace", index=False
    )
    md.generate_mapping_workbook(None, "Account", output_path, engine, "SourceAccounts", schema="dbo")

    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames.count("Account") == 1
    ws = wb["Account"]
    source_fields = {ws.cell(row=r, column=2).value for r in range(4, 6)}
    assert source_fields == {"Name", "City"}
