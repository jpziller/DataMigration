"""Coverage for reconciliation.py (roadmap #64) against a real SQLite
engine, built on real bulk_op() runs (same "real engine, real write
path" philosophy as test_bulkops_sqlite_integration.py).
"""
import openpyxl
import pandas as pd
import pytest

import bulkops as bo
import reconciliation as rc
import sql_client
from config import Settings
from stub_salesforce import StubBulkHandler, StubSF, describe_fields

_FIELDS = describe_fields(["Name", "LegacyId__c"])


def _stub_sf(handler):
    return StubSF({"Account": _FIELDS}, {"Account": handler})


@pytest.fixture
def sqlite_engine(tmp_path):
    s = Settings(
        sql_backend="sqlite",
        sql_sqlite_dir=str(tmp_path / "_sqlite"),
        sql_sqlite_schemas="dbo",
    )
    return sql_client.make_engine(s), s


def _write_mapping_workbook_with_source(path, object_name, source_table):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(object_name)
    ws.cell(row=1, column=1, value="Source Object:")
    ws.cell(row=1, column=2, value=source_table)
    wb.save(path)


def test_reconcile_flags_missing_load_table(sqlite_engine):
    engine, _ = sqlite_engine
    result = rc.reconcile_load_counts(engine, ["Account"], schema="dbo")
    assert result[0]["load_count"] is None
    assert any("doesn't exist yet" in f for f in result[0]["flags"])


def test_reconcile_flags_load_table_smaller_than_source(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    pd.DataFrame({"x": range(10)}).to_sql("SourceAccounts", engine, schema="dbo", index=False)
    pd.DataFrame({"LoadId": range(7), "LegacyId__c": [f"A{i}" for i in range(7)]}).to_sql(
        "Account_Load", engine, schema="dbo", index=False
    )
    mapping_path = str(tmp_path / "mapping.xlsx")
    _write_mapping_workbook_with_source(mapping_path, "Account", "SourceAccounts")

    result = rc.reconcile_load_counts(engine, ["Account"], schema="dbo", mapping_path=mapping_path)
    r = result[0]
    assert r["source_count"] == 10
    assert r["load_count"] == 7
    assert any("fewer rows than source" in f and "dropped 3 row(s)" in f for f in r["flags"])


def test_reconcile_clean_when_source_load_and_bulkops_all_agree(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    pd.DataFrame({"x": range(2)}).to_sql("SourceAccounts", engine, schema="dbo", index=False)
    df = pd.DataFrame({"LoadId": [1, 2], "LegacyId__c": ["A1", "A2"], "Name": ["Row1", "Row2"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    mapping_path = str(tmp_path / "mapping.xlsx")
    _write_mapping_workbook_with_source(mapping_path, "Account", "SourceAccounts")

    bo.enable_bulkops_logging(engine, schema="dbo")
    sf = _stub_sf(StubBulkHandler("LegacyId__c,Name,sf__Id\nA1,Row1,001X\nA2,Row2,001Y\n", ""))
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    result = rc.reconcile_load_counts(engine, ["Account"], schema="dbo", mapping_path=mapping_path)
    r = result[0]
    assert r["source_count"] == 2
    assert r["load_count"] == 2
    assert r["bulkops_submitted"] == 2
    assert r["bulkops_succeeded"] == 2
    assert r["flags"] == []


def test_reconcile_flags_never_loaded(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    pd.DataFrame({"LoadId": [1], "LegacyId__c": ["A1"]}).to_sql("Account_Load", engine, schema="dbo", index=False)

    result = rc.reconcile_load_counts(engine, ["Account"], schema="dbo")
    assert any("Never loaded via bulkops" in f for f in result[0]["flags"])


def test_reconcile_flags_stale_run_when_load_table_grew_since_last_bulkops(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1], "LegacyId__c": ["A1"], "Name": ["Row1"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)

    bo.enable_bulkops_logging(engine, schema="dbo")
    sf = _stub_sf(StubBulkHandler("LegacyId__c,Name,sf__Id\nA1,Row1,001X\n", ""))
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    # Load table grows after the bulkops run finished -- e.g. a rebuild.
    df2 = pd.DataFrame({"LoadId": [1, 2], "LegacyId__c": ["A1", "A2"], "Name": ["Row1", "Row2"], "Id": [None, None], "Error": [None, None]})
    df2.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)

    result = rc.reconcile_load_counts(engine, ["Account"], schema="dbo")
    assert any("stale prior run" in f for f in result[0]["flags"])


def test_reconcile_multiple_objects_independent(sqlite_engine):
    engine, _ = sqlite_engine
    pd.DataFrame({"LoadId": [1]}).to_sql("Account_Load", engine, schema="dbo", index=False)
    result = rc.reconcile_load_counts(engine, ["Account", "Contact"], schema="dbo")
    assert result[0]["object"] == "Account"
    assert result[0]["load_count"] == 1
    assert result[1]["object"] == "Contact"
    assert result[1]["load_count"] is None
