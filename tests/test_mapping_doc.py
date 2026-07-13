import pytest
import openpyxl

from mapping_doc import _safe_sheet_name, check_mapping_balance, extract_insert_columns, set_transform_script

SQL = """
-- staging
INSERT INTO dbo.Account_Load (Name, BillingCity, Legacy_Id__c)
SELECT Name, City, LegacyId FROM SourceAccounts;

INSERT INTO [dbo].[Contact_Load] ([FirstName], [LastName])
SELECT FirstName, LastName FROM SourceContacts;
"""


def test_extract_insert_columns_matches_named_table():
    cols = extract_insert_columns(SQL, "Account_Load")
    assert cols == ["Name", "BillingCity", "Legacy_Id__c"]


def test_extract_insert_columns_is_case_insensitive_on_table_name():
    cols = extract_insert_columns(SQL, "account_load")
    assert cols == ["Name", "BillingCity", "Legacy_Id__c"]


def test_extract_insert_columns_strips_brackets():
    cols = extract_insert_columns(SQL, "Contact_Load")
    assert cols == ["FirstName", "LastName"]


def test_extract_insert_columns_defaults_to_first_insert_when_table_none():
    cols = extract_insert_columns(SQL, None)
    assert cols == ["Name", "BillingCity", "Legacy_Id__c"]


def test_extract_insert_columns_returns_none_when_no_match():
    assert extract_insert_columns(SQL, "Opportunity_Load") is None


def test_extract_insert_columns_matches_sqlite_double_quoted_table():
    """SqliteDialect.qualify() produces "schema"."table" -- a real SQLite
    transform script using its own dialect's quoting used to never match
    here at all (found in review, _INSERT_INTO_RE only recognized bracket
    or bare identifiers)."""
    sqlite_sql = (
        'INSERT INTO "dbo"."Account_Load" ("Name", "BillingCity")\n'
        'SELECT Name, City FROM SourceAccounts;'
    )
    cols = extract_insert_columns(sqlite_sql, "Account_Load")
    assert cols == ["Name", "BillingCity"]


def test_extract_insert_columns_matches_sqlite_bare_double_quoted_table_no_schema():
    sqlite_sql = 'INSERT INTO "Account_Load" ("Name")\nSELECT Name FROM SourceAccounts;'
    cols = extract_insert_columns(sqlite_sql, "Account_Load")
    assert cols == ["Name"]


def test_extract_insert_columns_matches_mssql_select_into():
    # Found via a real dogfood run: sql_dialect.py's own
    # MssqlDialect.create_table_as_select_sql() -- this project's actual
    # canonical mssql *_Load-building pattern -- was never recognized at
    # all; check-mapping-balance raised "No INSERT INTO statement found"
    # against every one of this project's own real, working scripts.
    sql = (
        "DROP TABLE IF EXISTS [dbo].[Account_Load];\n\n"
        "SELECT\n"
        "    _MockRowId AS LoadId,\n"
        "    CAST(_MockRowId AS NVARCHAR(50)) AS MigrationID__c,\n"
        "    Name,\n"
        "    Type\n"
        "INTO [dbo].[Account_Load]\n"
        "FROM [dbo].[Account_Mock];\n"
    )
    cols = extract_insert_columns(sql, "Account_Load")
    assert cols == ["LoadId", "MigrationID__c", "Name", "Type"]


def test_extract_insert_columns_matches_mssql_select_into_qualified_source_columns():
    # A qualified column with no alias (e.g. joined child table "m.Subject")
    # must extract just the bare column name, matching what actually lands
    # in the target table.
    sql = (
        "SELECT\n"
        "    m._MockRowId AS LoadId,\n"
        "    m.Subject,\n"
        "    a.Id AS AccountId\n"
        "INTO [dbo].[Contact_Load]\n"
        "FROM [dbo].[Contact_Mock] m JOIN [dbo].[Account_Load] a ON a.LoadId = m._ParentMockRef;\n"
    )
    cols = extract_insert_columns(sql, "Contact_Load")
    assert cols == ["LoadId", "Subject", "AccountId"]


def test_extract_insert_columns_matches_sqlite_create_table_as_select():
    # The sqlite-backend equivalent of the mssql SELECT...INTO form above --
    # sql_dialect.py's SqliteDialect.create_table_as_select_sql().
    sql = (
        'CREATE TABLE "dbo"."Account_Load" AS\n'
        "SELECT\n"
        "    _MockRowId AS LoadId,\n"
        '    CAST(_MockRowId AS TEXT) AS "MigrationID__c",\n'
        "    Name\n"
        'FROM "dbo"."Account_Mock";\n'
    )
    cols = extract_insert_columns(sql, "Account_Load")
    assert cols == ["LoadId", "MigrationID__c", "Name"]


def test_extract_insert_columns_ignores_prose_in_header_comment_mentioning_select_into():
    # Found via a real dogfood run -- a genuinely embarrassing bug: this
    # project's own real 010_account_load.sql header comment describes the
    # SELECT...INTO port in English prose ("SELECT ... INTO is the
    # equivalent"), and the unstripped-comment regex matched that PROSE as
    # if it were real SQL (extracting "is" as the table name and comment
    # text as the column list) before comment-stripping was added.
    sql = (
        "/* Ported to real T-SQL -- CREATE TABLE ... AS SELECT isn't valid;\n"
        "   SELECT ... INTO is the equivalent. */\n\n"
        "DROP TABLE IF EXISTS [dbo].[Account_Load];\n\n"
        "SELECT\n"
        "    Name,\n"
        "    Type\n"
        "INTO [dbo].[Account_Load]\n"
        "FROM [dbo].[Account_Mock];\n"
    )
    cols = extract_insert_columns(sql, "Account_Load")
    assert cols == ["Name", "Type"]


def test_extract_insert_columns_ignores_line_comment_mentioning_insert_into():
    sql = (
        "-- old approach used INSERT INTO Foo (a, b) VALUES (...), replaced below\n"
        "SELECT\n"
        "    Name\n"
        "INTO [dbo].[Account_Load]\n"
        "FROM [dbo].[Account_Mock];\n"
    )
    cols = extract_insert_columns(sql, "Account_Load")
    assert cols == ["Name"]


def test_extract_insert_columns_matches_case_expression_alias():
    # 040_task_load.sql's real pattern: a multi-line CASE resolving a
    # polymorphic field, aliased at the very end -- the alias regex must
    # anchor to the end of the expression, not match an early false
    # positive inside the CASE itself.
    sql = (
        "SELECT\n"
        "    CASE m._ParentType\n"
        "        WHEN 'Account' THEN acc_direct.Id\n"
        "        WHEN 'Opportunity' THEN opp.Id\n"
        "    END AS WhatId\n"
        "INTO [dbo].[Task_Load]\n"
        "FROM [dbo].[Task_Mock] m;\n"
    )
    cols = extract_insert_columns(sql, "Task_Load")
    assert cols == ["WhatId"]


def test_safe_sheet_name_strips_invalid_excel_characters():
    assert _safe_sheet_name("Account:Contact/Test") == "Account_Contact_Test"


def test_safe_sheet_name_truncates_to_31_chars():
    name = "A" * 40
    result = _safe_sheet_name(name)
    assert len(result) == 31


class _StubObject:
    def __init__(self, fields):
        self._fields = fields

    def describe(self):
        return {"fields": self._fields}


class _StubSF:
    def __init__(self, object_name, fields):
        setattr(self, object_name, _StubObject(fields))


def _write_mapping_workbook(path, object_name, rows):
    """rows: list of (source_field, migrate_data, target_field) starting
    at row 4, matching mapping_doc.py's real column layout (source field
    API = col 2, Migrate Data = col 9, Target Field API = col 15)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(object_name)
    for i, (source, migrate, target) in enumerate(rows, start=4):
        ws.cell(row=i, column=2, value=source)
        ws.cell(row=i, column=9, value=migrate)
        ws.cell(row=i, column=15, value=target)
    wb.save(path)


def test_check_mapping_balance_detects_duplicate_target_field_in_one_sheet(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [
        ("first_name", "Yes", "Name"),
        ("company_name", "Yes", "Name"),  # duplicate target -- hard rule 14
    ])
    sql_path = tmp_path / "transform.sql"
    sql_path.write_text("INSERT INTO Account_Load (Name) SELECT x FROM y;", encoding="utf-8")

    sf = _StubSF("Account", [{"name": "Name"}])
    result = check_mapping_balance(sf, str(mapping_path), "Account", str(sql_path))

    assert result["duplicate_target_fields"] == {"Name": ["first_name", "company_name"]}


def test_check_mapping_balance_no_duplicates_when_targets_distinct(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [
        ("first_name", "Yes", "Name"),
        ("city", "Yes", "BillingCity"),
    ])
    sql_path = tmp_path / "transform.sql"
    sql_path.write_text("INSERT INTO Account_Load (Name, BillingCity) SELECT x, y FROM z;", encoding="utf-8")

    sf = _StubSF("Account", [{"name": "Name"}, {"name": "BillingCity"}])
    result = check_mapping_balance(sf, str(mapping_path), "Account", str(sql_path))

    assert result["duplicate_target_fields"] == {}
    assert result["duplicate_implemented_columns"] == []


def test_check_mapping_balance_detects_duplicate_implemented_column(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [("first_name", "Yes", "Name")])
    sql_path = tmp_path / "transform.sql"
    # Name listed twice in one INSERT INTO -- would break the real SQL.
    sql_path.write_text("INSERT INTO Account_Load (Name, Name) SELECT x, y FROM z;", encoding="utf-8")

    sf = _StubSF("Account", [{"name": "Name"}])
    result = check_mapping_balance(sf, str(mapping_path), "Account", str(sql_path))

    assert result["duplicate_implemented_columns"] == ["Name"]


def test_set_transform_script_fills_header_field(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [("first_name", "Yes", "Name")])
    (tmp_path / "sql" / "transformations").mkdir(parents=True)
    (tmp_path / "sql" / "transformations" / "010_account_load.sql").write_text("", encoding="utf-8")

    filename = set_transform_script(str(mapping_path), "Account", repo_root=str(tmp_path))

    assert filename == "010_account_load.sql"
    wb = openpyxl.load_workbook(mapping_path)
    ws = wb["Account"]
    assert ws.cell(row=1, column=5).value == "Transform Script:"
    assert ws.cell(row=1, column=6).value == "010_account_load.sql"


def test_set_transform_script_prefers_highest_numbered_match(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [("first_name", "Yes", "Name")])
    (tmp_path / "sql" / "transformations").mkdir(parents=True)
    (tmp_path / "sql" / "transformations" / "010_account_load.sql").write_text("", encoding="utf-8")
    (tmp_path / "sql" / "transformations" / "040_account_load.sql").write_text("", encoding="utf-8")

    filename = set_transform_script(str(mapping_path), "Account", repo_root=str(tmp_path))

    assert filename == "040_account_load.sql"


def test_set_transform_script_uses_source_ingestion_dir_when_given(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [("first_name", "Yes", "Name")])
    (tmp_path / "sql" / "source_ingestion").mkdir(parents=True)
    (tmp_path / "sql" / "source_ingestion" / "010_account_ingest.sql").write_text("", encoding="utf-8")

    filename = set_transform_script(
        str(mapping_path), "Account", script_subdir="source_ingestion", repo_root=str(tmp_path)
    )

    assert filename == "010_account_ingest.sql"


def test_set_transform_script_raises_when_no_sheet(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [("first_name", "Yes", "Name")])

    with pytest.raises(ValueError, match="No sheet named"):
        set_transform_script(str(mapping_path), "Contact", repo_root=str(tmp_path))


def test_set_transform_script_raises_when_no_matching_script(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [("first_name", "Yes", "Name")])
    (tmp_path / "sql" / "transformations").mkdir(parents=True)

    with pytest.raises(ValueError, match="No transform script"):
        set_transform_script(str(mapping_path), "Account", repo_root=str(tmp_path))


def test_set_transform_script_never_overwrites_source_object_header(tmp_path):
    mapping_path = tmp_path / "mapping.xlsx"
    _write_mapping_workbook(mapping_path, "Account", [("first_name", "Yes", "Name")])
    wb = openpyxl.load_workbook(mapping_path)
    ws = wb["Account"]
    ws.cell(row=1, column=1, value="Source Object:")
    ws.cell(row=1, column=2, value="SourceAccounts")
    wb.save(mapping_path)
    (tmp_path / "sql" / "transformations").mkdir(parents=True)
    (tmp_path / "sql" / "transformations" / "010_account_load.sql").write_text("", encoding="utf-8")

    set_transform_script(str(mapping_path), "Account", repo_root=str(tmp_path))

    wb2 = openpyxl.load_workbook(mapping_path)
    ws2 = wb2["Account"]
    assert ws2.cell(row=1, column=1).value == "Source Object:"
    assert ws2.cell(row=1, column=2).value == "SourceAccounts"
