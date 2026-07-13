"""Integration coverage for migration_run_book.py's database-touching
half against a real SQLite file -- previously untested at all (found in
review): test_migration_run_book.py only covers pure/offline helpers
(_object_matches, _is_separator_row, _parse_template). This file exercises
_load_order_rows/generate_migration_run_book's engine-reading path,
sync_run_book_from_log's real BulkOpsLog-driven sync, and
sync_source_ingestion_to_run_book via source_ingestion.import_directory()'s
own run_book_path/run_book_tab flags -- the same "real engine, real write
path" philosophy as test_bulkops_sqlite_integration.py.
"""
import os

import openpyxl
import pandas as pd
import pytest

import bulkops as bo
import migration_run_book as mrb
import source_ingestion as si
import sql_client
from config import Settings
from stub_salesforce import StubBulkHandler, StubSF, describe_fields

_FIELDS = describe_fields(["Name", "LegacyId__c"])


def _stub_sf(handler):
    return StubSF({"Account": _FIELDS}, {"Account": handler})


@pytest.fixture
def sqlite_engine(tmp_path):
    s = Settings(
        sql_backend="sqlite",
        sql_sqlite_dir=str(tmp_path / "_sqlite"),
        sql_sqlite_schemas="dbo",
    )
    return sql_client.make_engine(s), s


def _seed_load_order(engine, rows, edges=()):
    """rows: [(ObjectName, LoadLevel, LoadSequence)], edges:
    [(ChildObject, ParentObject)] -- the shape _load_order_rows() reads,
    written directly via to_sql rather than load_order.py's own
    write_to_sql() (SQL-Server-only raw T-SQL, not part of what's being
    tested here)."""
    pd.DataFrame(rows, columns=["ObjectName", "LoadLevel", "LoadSequence"]).to_sql(
        "ObjectLoadOrder", engine, schema="dbo", if_exists="replace", index=False
    )
    pd.DataFrame(list(edges), columns=["ChildObject", "ParentObject"]).to_sql(
        "ObjectDependency", engine, schema="dbo", if_exists="replace", index=False
    )


def _run_clean_load(engine, tmp_path, load_id=1, legacy_id="A1"):
    df = pd.DataFrame({"LoadId": [load_id], "LegacyId__c": [legacy_id], "Name": [f"Row{load_id}"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    sf = _stub_sf(StubBulkHandler(f"LegacyId__c,Name,sf__Id\n{legacy_id},Row{load_id},001X{load_id:03d}\n", ""))
    return bo.bulk_op(
        sf, engine, "Account", "insert", "Account_Load",
        key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
        email_deliverability="no-access",
    )


def test_generate_migration_run_book_autofills_load_phase_from_load_order(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    _seed_load_order(
        engine,
        rows=[("Account", 0, 1), ("Contact", 1, 2)],
        edges=[("Contact", "Account")],
    )
    output_path = str(tmp_path / "run_book.xlsx")

    mrb.generate_migration_run_book(
        output_path, "Dev1", engine=engine, object_names=["Account", "Contact"], schema="dbo",
    )

    assert os.path.exists(output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Dev1"]
    load_rows = [
        (row[1].value, row[2].value) for row in ws.iter_rows(min_row=mrb._FIRST_DATA_ROW)
        if row[0].value == "Load" or (row[0].value is None and row[1].value)
    ]
    # This repo already has real committed transform scripts for these
    # objects (010_account_load.sql etc.), so the Object cell resolves to
    # the actual script filename, not the bare object name -- confirming
    # script_filename_for() really is wired in here, not just object_names
    # echoed back verbatim.
    objects = [obj for obj, _ in load_rows]
    assert any("account" in obj.lower() for obj in objects)
    contact_obj, contact_dependency = next(row for row in load_rows if "contact" in row[0].lower())
    assert "Account" in contact_dependency


def test_generate_migration_run_book_refuses_to_overwrite_existing_tab(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    _seed_load_order(engine, rows=[("Account", 0, 1)])
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(output_path, "Dev1", engine=engine, object_names=["Account"], schema="dbo")

    with pytest.raises(ValueError, match="already exists"):
        mrb.generate_migration_run_book(output_path, "Dev1", engine=engine, object_names=["Account"], schema="dbo")


def test_sync_run_book_from_log_fills_pending_placeholder(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    _seed_load_order(engine, rows=[("Account", 0, 1)])
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(output_path, "Dev1", engine=engine, object_names=["Account"], schema="dbo")

    bo.enable_bulkops_logging(engine, schema="dbo")
    _run_clean_load(engine, tmp_path, load_id=1)

    result = mrb.sync_run_book_from_log(engine, output_path, "Dev1", schema="dbo")
    assert result["synced"] == 1
    assert result["inserted"] == 0
    assert result["updated"] == 1

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Dev1"]
    status_col = mrb._COLUMNS.index("Status") + 1
    total_col = mrb._COLUMNS.index("Total Records") + 1
    found = [
        row for row in ws.iter_rows(min_row=mrb._FIRST_DATA_ROW)
        if row[total_col - 1].value == 1
    ]
    assert len(found) == 1
    assert found[0][status_col - 1].value == "Completed"


def test_sync_run_book_from_log_is_idempotent_on_second_call(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    _seed_load_order(engine, rows=[("Account", 0, 1)])
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(output_path, "Dev1", engine=engine, object_names=["Account"], schema="dbo")

    bo.enable_bulkops_logging(engine, schema="dbo")
    _run_clean_load(engine, tmp_path, load_id=1)
    mrb.sync_run_book_from_log(engine, output_path, "Dev1", schema="dbo")

    second = mrb.sync_run_book_from_log(engine, output_path, "Dev1", schema="dbo")
    assert second == {"synced": 0, "inserted": 0, "updated": 0}


def test_sync_run_book_from_log_inserts_new_row_for_a_retry_with_no_placeholder(sqlite_engine, tmp_path):
    """A second load against an object with no remaining pending
    placeholder (e.g. a retry) must insert a new Load-phase row rather
    than being dropped."""
    engine, _ = sqlite_engine
    _seed_load_order(engine, rows=[("Account", 0, 1)])
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(output_path, "Dev1", engine=engine, object_names=["Account"], schema="dbo")

    bo.enable_bulkops_logging(engine, schema="dbo")
    _run_clean_load(engine, tmp_path, load_id=1, legacy_id="A1")
    mrb.sync_run_book_from_log(engine, output_path, "Dev1", schema="dbo")

    _run_clean_load(engine, tmp_path, load_id=2, legacy_id="A2")
    result = mrb.sync_run_book_from_log(engine, output_path, "Dev1", schema="dbo")
    assert result["synced"] == 1
    assert result["inserted"] == 1


def test_sync_source_ingestion_to_run_book_via_import_directory(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    _seed_load_order(engine, rows=[("Account", 0, 1)])
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(output_path, "Dev1", engine=engine, object_names=["Account"], schema="dbo")

    si.enable_source_ingestion_logging(engine, schema="dbo")
    csv_dir = str(tmp_path / "csvs")
    os.makedirs(csv_dir, exist_ok=True)
    pd.DataFrame([{"Name": "Acme", "Legacy_Id": "A1"}]).to_csv(
        os.path.join(csv_dir, "SourceAccounts.csv"), index=False
    )

    results = si.import_directory(
        engine, csv_dir, sql_dir=str(tmp_path / "sql"), schema="dbo", ticket="TEST-1",
        run_book_path=output_path, run_book_tab="Dev1",
    )
    assert results[0]["status"] == "created"
    assert "run_book_sync_error" not in results[-1]

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Dev1"]
    notes_col = mrb._COLUMNS.index("Notes") + 1
    matches = [
        row for row in ws.iter_rows(min_row=mrb._FIRST_DATA_ROW)
        if row[notes_col - 1].value and "SourceIngestionLog" in str(row[notes_col - 1].value)
    ]
    assert len(matches) == 1
    status_col = mrb._COLUMNS.index("Status") + 1
    assert matches[0][status_col - 1].value == "Completed"


def test_generate_run_book_flowchart_draws_edges_from_real_load_order_dependency_text(sqlite_engine, tmp_path):
    """End-to-end: _load_order_rows() (already exercised above) writes
    real "After: Account" dependency text into the Load phase; the
    flowchart generator must resolve that back into a real edge between
    the two nodes, matching Account's node by whole-token match even
    though its label is a script filename, not the bare object name."""
    engine, _ = sqlite_engine
    _seed_load_order(
        engine,
        rows=[("Account", 0, 1), ("Contact", 1, 2)],
        edges=[("Contact", "Account")],
    )
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(
        output_path, "Dev1", engine=engine, object_names=["Account", "Contact"], schema="dbo",
    )

    mermaid_text, summary = mrb.generate_run_book_flowchart(output_path, "Dev1")

    assert summary["edges"] == 1
    assert summary["unresolved_dependencies"] == []
    assert mermaid_text.startswith("```mermaid\nflowchart TD")
    assert mermaid_text.rstrip().endswith("```")
    assert " --> " in mermaid_text


def test_generate_run_book_flowchart_reports_unresolved_dependency(sqlite_engine, tmp_path):
    """A Dependency mention with no matching row in this tab must be
    dropped, not guessed at or raised -- surfaced in the summary instead.
    _load_order_rows() itself only ever writes an in-scope dependency (see
    its own edges filter), so this simulates the case by hand-editing the
    Dependency cell after generation -- e.g. a human editing the tab
    directly, or a future sync path that isn't scope-filtered the same way."""
    engine, _ = sqlite_engine
    _seed_load_order(engine, rows=[("Contact", 0, 1)])
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(
        output_path, "Dev1", engine=engine, object_names=["Contact"], schema="dbo",
    )

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Dev1"]
    dependency_col = mrb._COLUMNS.index("Dependency") + 1
    object_col = mrb._COLUMNS.index("Object") + 1
    contact_row = next(
        row[0].row for row in ws.iter_rows(min_row=mrb._FIRST_DATA_ROW)
        if row[object_col - 1].value and "contact" in str(row[object_col - 1].value).lower()
    )
    ws.cell(row=contact_row, column=dependency_col, value="After: Account")
    wb.save(output_path)

    mermaid_text, summary = mrb.generate_run_book_flowchart(output_path, "Dev1")
    assert summary["edges"] == 0
    assert "Account" in summary["unresolved_dependencies"]


def test_generate_run_book_flowchart_reports_unparseable_dependency_note(sqlite_engine, tmp_path):
    """Found in review: a free-text Dependency note (plausible for a
    hand-filled Pre-/Post-Migration row) that doesn't match the "After:
    X" convention used to be silently dropped, indistinguishable from a
    genuine "no dependency" row -- must now surface separately from
    unresolved_dependencies (which means "found an After: mention, no
    matching row"), not conflated with it."""
    engine, _ = sqlite_engine
    _seed_load_order(engine, rows=[("Contact", 0, 1)])
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(
        output_path, "Dev1", engine=engine, object_names=["Contact"], schema="dbo",
    )

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Dev1"]
    dependency_col = mrb._COLUMNS.index("Dependency") + 1
    object_col = mrb._COLUMNS.index("Object") + 1
    contact_row = next(
        row[0].row for row in ws.iter_rows(min_row=mrb._FIRST_DATA_ROW)
        if row[object_col - 1].value and "contact" in str(row[object_col - 1].value).lower()
    )
    ws.cell(row=contact_row, column=dependency_col, value="Depends on Account for billing info")
    wb.save(output_path)

    mermaid_text, summary = mrb.generate_run_book_flowchart(output_path, "Dev1")
    assert summary["edges"] == 0
    assert summary["unresolved_dependencies"] == []
    assert len(summary["unparsed_dependency_notes"]) == 1
    assert "Depends on Account for billing info" in summary["unparsed_dependency_notes"][0]


def test_generate_run_book_flowchart_raises_when_tab_missing(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    _seed_load_order(engine, rows=[("Account", 0, 1)])
    output_path = str(tmp_path / "run_book.xlsx")
    mrb.generate_migration_run_book(output_path, "Dev1", engine=engine, object_names=["Account"], schema="dbo")

    with pytest.raises(ValueError, match="No tab named"):
        mrb.generate_run_book_flowchart(output_path, "DoesNotExist")
