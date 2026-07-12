"""Coverage for failure_triage.py (roadmap #61) against a real SQLite
engine -- a completed bulk_op() run's failures written back for real,
then triaged, same "real engine, real write path" philosophy as
test_bulkops_sqlite_integration.py.
"""
import openpyxl
import pandas as pd
import pytest

import bulkops as bo
import failure_triage as ft
import sql_client
from config import Settings
from stub_salesforce import StubBulkHandler, StubSF, describe_fields

_FIELDS = describe_fields(["Name", "LegacyId__c"])


def _stub_sf(handler):
    return StubSF({"Account": _FIELDS}, {"Account": handler})


@pytest.fixture
def sqlite_engine(tmp_path):
    s = Settings(
        sql_backend="sqlite",
        sql_sqlite_dir=str(tmp_path / "_sqlite"),
        sql_sqlite_schemas="dbo",
    )
    return sql_client.make_engine(s), s


def _write_mapping_workbook(path, object_name, rows):
    """rows: list of (source_field, migrate_data, target_field), matching
    mapping_doc.py's real column layout (source field API = col 2,
    Migrate Data = col 9, Target Field API = col 15)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(object_name)
    for i, (source, migrate, target) in enumerate(rows, start=4):
        ws.cell(row=i, column=2, value=source)
        ws.cell(row=i, column=9, value=migrate)
        ws.cell(row=i, column=15, value=target)
    wb.save(path)


def test_triage_raises_when_no_error_column(sqlite_engine):
    engine, _ = sqlite_engine
    pd.DataFrame({"LoadId": [1]}).to_sql("Account_Load", engine, schema="dbo", index=False)
    with pytest.raises(ValueError, match="has bulkops been run"):
        ft.triage_failures(engine, "Account_Load", schema="dbo")


def test_triage_empty_when_no_failures(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1], "LegacyId__c": ["A1"], "Name": ["Row1"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    sf = _stub_sf(StubBulkHandler("LegacyId__c,Name,sf__Id\nA1,Row1,001X\n", ""))
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    assert ft.triage_failures(engine, "Account_Load", schema="dbo") == []


def test_triage_groups_by_normalized_signature_with_known_code(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1, 2, 3], "LegacyId__c": ["A1", "A2", "A3"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)

    fields = describe_fields(["LegacyId__c"])
    fail_csv = (
        "LegacyId__c,sf__Error\n"
        "A1,DUPLICATE_VALUE:duplicate value found: record with id: 001XX000003DHPbYAO\n"
        "A2,DUPLICATE_VALUE:duplicate value found: record with id: 001XX000003DHQcYAO\n"
        "A3,MALFORMED_ID:bad reference\n"
    )
    sf = StubSF({"Account": fields}, {"Account": StubBulkHandler("", fail_csv)})
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    results = ft.triage_failures(engine, "Account_Load", schema="dbo")
    assert len(results) == 2
    # Sorted by count descending -- the 2-row DUPLICATE_VALUE group first.
    assert results[0]["code"] == "DUPLICATE_VALUE"
    assert results[0]["count"] == 2
    assert "duplicate rule" in results[0]["cause"].lower()
    assert results[1]["code"] == "MALFORMED_ID"
    assert results[1]["count"] == 1


def test_triage_unrecognized_code_falls_back_gracefully(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1], "LegacyId__c": ["A1"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    fields = describe_fields(["LegacyId__c"])
    fail_csv = "LegacyId__c,sf__Error\nA1,SOME_BRAND_NEW_CODE:never seen before\n"
    sf = StubSF({"Account": fields}, {"Account": StubBulkHandler("", fail_csv)})
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    results = ft.triage_failures(engine, "Account_Load", schema="dbo")
    assert len(results) == 1
    assert results[0]["code"] == "SOME_BRAND_NEW_CODE"
    assert "no known guidance" in results[0]["cause"].lower()
    assert results[0]["next_steps"]


def test_triage_extracts_required_field_names(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1], "LegacyId__c": ["A1"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    fields = describe_fields(["LegacyId__c"])
    fail_csv = 'LegacyId__c,sf__Error\nA1,"REQUIRED_FIELD_MISSING:Required fields are missing: [BillingCity, BillingState]"\n'
    sf = StubSF({"Account": fields}, {"Account": StubBulkHandler("", fail_csv)})
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    results = ft.triage_failures(engine, "Account_Load", schema="dbo")
    assert results[0]["fields"] == ["BillingCity", "BillingState"]
    assert results[0]["detail"] == []  # no mapping_path/object_name given -- no enrichment


def test_triage_required_field_never_mapped(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1], "LegacyId__c": ["A1"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    fields = describe_fields(["LegacyId__c"])
    fail_csv = "LegacyId__c,sf__Error\nA1,REQUIRED_FIELD_MISSING:Required fields are missing: [BillingCity]\n"
    sf = StubSF({"Account": fields}, {"Account": StubBulkHandler("", fail_csv)})
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    mapping_path = str(tmp_path / "mapping.xlsx")
    _write_mapping_workbook(mapping_path, "Account", [
        ("Name", "Yes", "Name"),
        ("Legacy_Id", "Yes", "Legacy_Id__c"),
    ])

    results = ft.triage_failures(engine, "Account_Load", schema="dbo",
                                  object_name="Account", mapping_path=mapping_path)
    assert any("never chosen as a Target Field" in d for d in results[0]["detail"])


def test_triage_required_field_already_mapped(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1], "LegacyId__c": ["A1"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    fields = describe_fields(["LegacyId__c"])
    fail_csv = "LegacyId__c,sf__Error\nA1,REQUIRED_FIELD_MISSING:Required fields are missing: [BillingCity]\n"
    sf = StubSF({"Account": fields}, {"Account": StubBulkHandler("", fail_csv)})
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    mapping_path = str(tmp_path / "mapping.xlsx")
    _write_mapping_workbook(mapping_path, "Account", [
        ("City", "Yes", "BillingCity"),
    ])

    results = ft.triage_failures(engine, "Account_Load", schema="dbo",
                                  object_name="Account", mapping_path=mapping_path)
    assert any("already mapped" in d for d in results[0]["detail"])


def test_triage_cross_references_validation_rule_candidates(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1], "LegacyId__c": ["A1"]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    fields = describe_fields(["LegacyId__c"])
    fail_csv = "LegacyId__c,sf__Error\nA1,FIELD_CUSTOM_VALIDATION_EXCEPTION:Name is required\n"
    sf = StubSF({"Account": fields}, {"Account": StubBulkHandler("", fail_csv)})
    bo.bulk_op(sf, engine, "Account", "insert", "Account_Load",
               key_column="LoadId", schema="dbo", stage_dir=str(tmp_path / "_stage"),
               email_deliverability="no-access")

    pd.DataFrame([
        {"ObjectName": "Account", "CheckType": "ValidationRule", "ItemName": "No_Blank_Name",
         "IsActive": 1, "DirectHit": 0, "Detail": "Name is required", "AnalyzedDate": "2026-07-12"},
        {"ObjectName": "Account", "CheckType": "ValidationRule", "ItemName": "Inactive_Rule",
         "IsActive": 0, "DirectHit": 0, "Detail": "Should never show up", "AnalyzedDate": "2026-07-12"},
    ]).to_sql("ObjectAutomationRisk", engine, schema="dbo", if_exists="replace", index=False)

    results = ft.triage_failures(engine, "Account_Load", schema="dbo", object_name="Account")
    assert any("No_Blank_Name" in d for d in results[0]["detail"])
    assert not any("Inactive_Rule" in d for d in results[0]["detail"])


def test_triage_sorted_by_count_descending(sqlite_engine, tmp_path):
    engine, _ = sqlite_engine
    df = pd.DataFrame({"LoadId": [1, 2, 3, 4]})
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)
    df["Error"] = ["MALFORMED_ID:x"] * 1 + ["STRING_TOO_LONG:y"] * 3
    df.to_sql("Account_Load", engine, schema="dbo", if_exists="replace", index=False)

    results = ft.triage_failures(engine, "Account_Load", schema="dbo")
    assert [r["count"] for r in results] == [3, 1]
    assert results[0]["code"] == "STRING_TOO_LONG"
