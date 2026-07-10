"""Data model ERD diagrams -- source subject-area models + a target model
(roadmap #57), styled to approximate Salesforce Data Model Notation (SDMN)
in Mermaid, for import into Lucid.

Verified against developer.salesforce.com's actual SDMN guide before
building this: SDMN encodes real information in per-entity fill color,
border style (solid/dashed/dotted/none), and a diamond-vs-line-vs-curve
relationship symbol (master-detail/lookup/recursive).

Renders as Mermaid `classDiagram`, not `erDiagram` or `flowchart` --
confirmed against Mermaid's own current docs to be the one diagram type
that gets all three SDMN-relevant features at once:
  - UML composition (`*--`, filled diamond) vs aggregation (`o--`, hollow
    diamond) for master-detail vs lookup -- a genuinely closer visual match
    to SDMN's own diamond-on-the-parent-side convention than erDiagram's
    solid/dashed lines or flowchart's thin/thick arrows.
  - Real per-class fill/border color via `classDef`/`:::` styling --
    something erDiagram cannot do at all. Palette reused from
    forcedotcom/sf-skills' own `external-diagram-mermaid-generate` skill
    (Standard=sky blue, Custom=orange, External=green) rather than
    inventing a new one -- their skill already validated this pattern,
    no reason to pick different colors for the same STD/CUST/EXT axis.
  - Attribute lists per class (PK/FK marked as trailing text -- Mermaid
    has no native key-stereotype syntax for individual attributes) plus
    `"1" *-- "1..*"`-style cardinality strings on each relationship end.

Target model: fully automatable. load_order.build_dependency_edges()
already returns real, describe()-driven relationships
(is_master_detail/is_nillable) -- reused directly, not reimplemented.
Object-type color coding (Standard/Custom/External) comes from describe()
itself (the `custom` flag, and the `__x` API-name suffix for external
objects) -- also real, not guessed.

Source model(s): staging tables carry no FK constraints and no
describe()-equivalent relationship metadata, so relationships here are a
naming-convention HEURISTIC ONLY, always labeled "(guessed)", always
rendered as the weaker aggregation form, and never colored by object type
(there's no Standard/Custom/External axis for a plain SQL table) --
matching this framework's "auto-map suggests, human decides" discipline
(hard rule 11's spirit). Subject-area grouping is an explicit, human-
supplied input, never auto-clustered -- there's no reliable signal to
cluster source tables on automatically.
"""
import re

import openpyxl
from sqlalchemy import text

from load_order import build_dependency_edges
from mapping_doc import _COL_MIGRATE_DATA, _COL_SOURCE_FIELD_API
from risk_analyzer import fields_in_scope_from_mapping

_FK_HEURISTIC_RE = re.compile(r"^(.+?)_?[Ii]d$")

# Reused verbatim from forcedotcom/sf-skills' external-diagram-mermaid-generate
# skill's own color palette -- an already-validated STD/CUST/EXT convention,
# not a new one invented here.
_OBJECT_TYPE_CLASS_DEFS = {
    "standardObject": "fill:#bae6fd,stroke:#0369a1,color:#1f2937",
    "customObject": "fill:#fed7aa,stroke:#c2410c,color:#1f2937",
    "externalObject": "fill:#a7f3d0,stroke:#047857,color:#1f2937",
}


def _object_type_css_class(object_name, is_custom):
    if object_name.endswith("__x"):
        return "externalObject"
    if is_custom:
        return "customObject"
    return "standardObject"


def _render_class(entity_name, attributes):
    lines = [f"    class {entity_name} {{"]
    for field_type, field_name, key_label in attributes:
        key_part = f" {key_label}" if key_label else ""
        lines.append(f"        +{field_type} {field_name}{key_part}")
    lines.append("    }")
    return "\n".join(lines)


def _render_relationship(parent, child, label, is_master_detail, is_nillable, guessed=False):
    """Mermaid classDiagram relationship: `Parent "1" *-- "1..*" Child : "label"`.
    Composition (*--, filled diamond) = master-detail; aggregation (o--,
    hollow diamond) = lookup -- SDMN itself uses a diamond specifically for
    master-detail, so this is a closer match than a plain line-style
    distinction. Child-side cardinality is "1..*" when the field is
    required (not nillable), "0..*" otherwise. guessed=True appends
    " (guessed)" to the label and always renders as aggregation, regardless
    of is_master_detail -- a guess should never visually claim the
    stronger, more confident relationship type."""
    arrow = "o--" if (guessed or not is_master_detail) else "*--"
    child_card = "1..*" if not is_nillable else "0..*"
    display_label = f"{label} (guessed)" if guessed else label
    return f'    {parent} "1" {arrow} "{child_card}" {child} : "{display_label}"'


def _wrap_diagram(title, class_blocks, relationship_lines, css_classes_by_entity=None, notes=None):
    parts = [f"# {title}", "", "```mermaid", "classDiagram"]

    used_defs = sorted(set((css_classes_by_entity or {}).values()))
    for css_class in used_defs:
        parts.append(f"    classDef {css_class} {_OBJECT_TYPE_CLASS_DEFS[css_class]}")
    if used_defs:
        parts.append("")

    parts.extend(class_blocks)

    if css_classes_by_entity:
        parts.append("")
        for entity_name, css_class in css_classes_by_entity.items():
            parts.append(f"    class {entity_name}:::{css_class}")

    if relationship_lines:
        parts.append("")
        parts.extend(relationship_lines)
    parts.append("```")
    if notes:
        parts.append("")
        parts.extend(notes)
    return "\n".join(parts) + "\n"


def generate_target_model_diagram(sf, object_names, mapping_path=None):
    """Mermaid classDiagram ERD for a target-org data model (core + custom
    objects). Relationships come straight from
    load_order.build_dependency_edges() -- real, describe()-driven, not
    guessed. Attributes default to Id/Name/required/reference fields;
    --mapping-path scopes each object's attribute list to whatever's
    actually flagged Migrate Data = Yes for it instead. Each object is
    colored by its real describe()-derived type (Standard/Custom/External).
    Returns the full .md file text (title + fenced mermaid block)."""
    scoped_fields = fields_in_scope_from_mapping(mapping_path, object_names) if mapping_path else {}

    class_blocks = []
    css_classes_by_entity = {}
    for object_name in object_names:
        desc = getattr(sf, object_name).describe()
        css_classes_by_entity[object_name] = _object_type_css_class(object_name, desc.get("custom", False))
        allowed = scoped_fields.get(object_name)
        attributes = []
        for f in desc["fields"]:
            name = f["name"]
            if allowed is not None:
                if name not in allowed and name != "Id":
                    continue
            else:
                is_structural = (
                    name in ("Id", "Name")
                    or not f.get("nillable", True)
                    or f["type"] == "reference"
                )
                if not is_structural:
                    continue
            key_label = "PK" if name == "Id" else ("FK" if f["type"] == "reference" else None)
            attributes.append((f["type"], name, key_label))
        class_blocks.append(_render_class(object_name, attributes))

    edges = build_dependency_edges(sf, object_names)
    relationship_lines = [
        _render_relationship(e["parent"], e["child"], e["field"], e["is_master_detail"], e["is_nillable"])
        for e in edges
    ]

    return _wrap_diagram(
        f"Target data model: {', '.join(object_names)}",
        class_blocks, relationship_lines, css_classes_by_entity=css_classes_by_entity,
    )


def _source_table_columns(engine, schema, table_name):
    with engine.connect() as cx:
        rows = cx.execute(
            text(
                "SELECT COLUMN_NAME, DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = :schema AND TABLE_NAME = :table ORDER BY ORDINAL_POSITION"
            ),
            {"schema": schema, "table": table_name},
        ).mappings().all()
    if not rows:
        raise ValueError(f"No such table: {schema}.{table_name}")
    return [(r["DATA_TYPE"], r["COLUMN_NAME"]) for r in rows]


def _source_table_primary_key_columns(engine, schema, table_name):
    """Real SQL Server PK constraint columns, if any -- unlike the FK
    heuristic, this is genuine ground truth, not a guess. Most staging
    tables won't have one; that's fine, it just means no column gets a
    PK label rather than a wrong guess."""
    with engine.connect() as cx:
        rows = cx.execute(
            text(
                "SELECT kcu.COLUMN_NAME FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc "
                "JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE kcu "
                "  ON tc.CONSTRAINT_NAME = kcu.CONSTRAINT_NAME AND tc.TABLE_SCHEMA = kcu.TABLE_SCHEMA "
                "WHERE tc.CONSTRAINT_TYPE = 'PRIMARY KEY' AND tc.TABLE_SCHEMA = :schema AND tc.TABLE_NAME = :table"
            ),
            {"schema": schema, "table": table_name},
        ).all()
    return {r[0] for r in rows}


def _source_fields_in_scope(mapping_path, table_name):
    """Source-side counterpart to risk_analyzer.fields_in_scope_from_mapping
    -- that function scopes by TARGET field; this scopes by SOURCE field,
    keyed off the mapping doc's "Source Object" column matching
    table_name. Returns None (no scoping) if no mapping doc is given."""
    if not mapping_path:
        return None
    wb = openpyxl.load_workbook(mapping_path, data_only=True)
    fields = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4):
            source_object = row[0].value
            if source_object is None or str(source_object).strip() != table_name:
                continue
            migrate = row[_COL_MIGRATE_DATA - 1].value
            source_field = row[_COL_SOURCE_FIELD_API - 1].value
            if migrate and str(migrate).strip() == "Yes" and source_field:
                fields.add(str(source_field).strip())
    return fields


def _guess_fk_relationships(table_names, columns_by_table):
    """Naming-heuristic only: a column matching <name>_?Id at the end
    proposes a relationship to another table in this same call whose name
    contains <name> -- e.g. AccountId in SourceContacts -> SourceAccounts.
    Never authoritative; always reported separately for human review."""
    guesses = []
    for table_name in table_names:
        for _dtype, col_name in columns_by_table[table_name]:
            m = _FK_HEURISTIC_RE.match(col_name)
            if not m:
                continue
            stem = m.group(1).lower()
            if not stem:
                continue
            for other_table in table_names:
                if other_table == table_name:
                    continue
                if stem in other_table.lower():
                    guesses.append({
                        "child": table_name, "parent": other_table, "field": col_name,
                    })
                    break
    return guesses


def generate_source_model_diagram(engine, table_names, schema="dbo", mapping_path=None):
    """Mermaid classDiagram ERD for one subject-area's source tables.
    Relationships are a naming-convention guess only (see
    _guess_fk_relationships) -- always rendered as aggregation (the
    weaker, non-owning relationship type) and labeled "(guessed)". No
    object-type color coding here -- Standard/Custom/External is a
    Salesforce-object concept with no equivalent for a plain SQL table.
    Returns (file_text, guessed_relationships) so the caller can print the
    guesses for explicit human review."""
    columns_by_table = {t: _source_table_columns(engine, schema, t) for t in table_names}
    pk_columns_by_table = {t: _source_table_primary_key_columns(engine, schema, t) for t in table_names}
    guesses = _guess_fk_relationships(table_names, columns_by_table)
    # Only a column that actually produced a guessed relationship is
    # labeled FK -- matching the FK naming pattern alone isn't enough
    # (found live: a table's own primary key, e.g. account_id, matches
    # the same "_id" pattern but isn't a foreign key at all).
    guessed_fk_columns_by_table = {t: set() for t in table_names}
    for g in guesses:
        guessed_fk_columns_by_table[g["child"]].add(g["field"])

    class_blocks = []
    for table_name in table_names:
        allowed = _source_fields_in_scope(mapping_path, table_name)
        attributes = []
        for dtype, col_name in columns_by_table[table_name]:
            if allowed is not None and col_name not in allowed:
                continue
            if col_name in pk_columns_by_table[table_name]:
                key_label = "PK"
            elif col_name in guessed_fk_columns_by_table[table_name]:
                key_label = "FK"
            else:
                key_label = None
            attributes.append((dtype, col_name, key_label))
        class_blocks.append(_render_class(table_name, attributes))

    relationship_lines = [
        _render_relationship(g["parent"], g["child"], g["field"],
                              is_master_detail=False, is_nillable=True, guessed=True)
        for g in guesses
    ]

    notes = [
        "> Relationships in this source model are a **naming-convention guess only** "
        "(no foreign keys exist on staging tables) -- verify every one before relying on it.",
    ]
    file_text = _wrap_diagram(f"Source data model: {', '.join(table_names)}", class_blocks, relationship_lines, notes=notes)
    return file_text, guesses
