"""Bulk-load failure triage assistant (roadmap #61).

Groups a completed bulk_op() run's failures by normalized error signature
(bulkops.py's own _normalize_error_signature() -- record-Id tokens
collapsed to <ID>, so the same recurring cause counts as one signature,
not one per row) and maps well-known, stable Salesforce Bulk API error
CODES to a likely root cause and which existing command to run next --
turning "100 rows failed" into "1 root cause, here's where to look"
instead of manually reading raw error strings row by row.

Advisory only, same "suggests, never auto-fixes" posture as
auto_mapper.py -- this never changes data, never re-runs bulkops, and
never claims certainty it doesn't have. Two disclosed scope limits, not
oversights:
  - Field-name extraction is only attempted for REQUIRED_FIELD_MISSING,
    whose "Required fields are missing: [Field1, Field2]" bracketed-list
    shape has been stable across Salesforce API versions -- every other
    error code gets guidance text only, since inventing a field-position
    regex for a message shape not directly confirmed against this
    project's own live data would be exactly the kind of stale-training-
    data guess CLAUDE.md warns against for anything API-version-specific.
  - DUPLICATE_VALUE gets guidance text only, never a live cross-reference
    against risk_analyzer.py's ObjectAutomationRisk data -- that table
    only ever scans ValidationRule/ApexTrigger/WorkflowRule/
    ApprovalProcess/FlowDefinitionView (see risk_analyzer.py's own
    docstring), never Salesforce's separate DuplicateRule metadata type,
    so there's genuinely nothing on file to cross-reference yet.
"""
import re

import openpyxl
from sqlalchemy import text

import sql_dialect
from bulkops import _normalize_error_signature
from mapping_doc import _COL_TARGET_FIELD_API, _safe_sheet_name

_REQUIRED_FIELD_MISSING_RE = re.compile(
    r"Required fields are missing:\s*\[([^\]]*)\]", re.IGNORECASE
)

# Cause + next-step guidance per well-documented Salesforce Bulk API error
# code (the part of an error message before its first ":"). Deliberately
# generic/command-pointing rather than trying to pinpoint an exact field
# or row from message text this tool can't reliably parse -- see this
# module's own docstring for the two places field-level detail IS
# attempted, and why those two are judged safe to attempt.
_ERROR_CODE_GUIDANCE = {
    "DUPLICATE_VALUE": {
        "cause": "A duplicate rule or a unique-field constraint rejected this row.",
        "next_steps": [
            "Confirm the migration key/external ID field genuinely has no duplicate values in the source -- check-load-table-duplicate-keys and validate-external-id should already catch this before a load; re-run them if this is showing up anyway.",
            "If the migration key is clean, this is likely a real Salesforce duplicate/matching rule -- these aren't captured by analyze-org-risk today (it scans ValidationRule/ApexTrigger/WorkflowRule/ApprovalProcess/FlowDefinitionView, not Salesforce's separate DuplicateRule metadata), so confirm directly in Setup for now.",
        ],
    },
    "REQUIRED_FIELD_MISSING": {
        "cause": "Salesforce rejected the row because a required field had no value.",
        "next_steps": [
            "Run check-required-mappings against this object's mapping doc -- was the named field ever chosen as a Target Field at all?",
            "If it WAS mapped, run profile-sql-table against the source table -- a required target field fed from a source field that isn't always populated will fail for exactly the unpopulated rows, not all of them.",
        ],
    },
    "STRING_TOO_LONG": {
        "cause": "A source value is longer than the target field's max length.",
        "next_steps": [
            "Cross-reference profile-sql-table's max observed length for the source field against the target field's real length from describe()/dump-describe.",
        ],
    },
    "INVALID_CROSS_REFERENCE_KEY": {
        "cause": "A lookup/master-detail field points at an Id that doesn't exist in the target org yet.",
        "next_steps": [
            "Almost always a load-order problem, not a data problem -- confirm analyze-load-order's sequence was followed and add-bulk-load-sort-column (the Parent-Batch Sort Rule) was applied before this load.",
            "If the sequence is correct, the source data may reference a parent record that was never brought into scope for this migration -- confirm with the client.",
        ],
    },
    "FIELD_CUSTOM_VALIDATION_EXCEPTION": {
        "cause": "An active validation rule's formula rejected this row.",
        "next_steps": [
            "Cross-reference analyze-org-risk's cached active validation rules for this object -- their ErrorMessage text often names the exact business rule that fired.",
        ],
    },
    "INVALID_FIELD_FOR_INSERT_UPDATE": {
        "cause": "A field in the request isn't createable/updateable for this operation.",
        "next_steps": [
            "bulk_op()'s own pre-flight check normally catches this before submission -- seeing it here anyway suggests field-level security changed after the pre-flight check last ran. Re-confirm profile/permission-set access (the Field-Level Security Bundling Rule, hard rule 8) for the field named in the error.",
        ],
    },
    "MALFORMED_ID": {
        "cause": "A field expecting a Salesforce Id received a value that isn't a valid Id shape.",
        "next_steps": [
            "Check the source/transform for this field -- likely a blank, truncated, or non-Id value (e.g. a legacy system's own key) being sent where a real Salesforce Id is expected.",
        ],
    },
    "UNABLE_TO_LOCK_ROW": {
        "cause": "Row-lock contention -- concurrent batches touched the same parent record.",
        "next_steps": [
            "Confirm add-bulk-load-sort-column (the Parent-Batch Sort Rule) was applied so same-parent rows land in the same batch. recommend-batch-size can also suggest a smaller batch size for this object.",
        ],
    },
}


def _extract_required_fields(message):
    """Best-effort extraction of the field name(s) named in a
    REQUIRED_FIELD_MISSING error's "Required fields are missing:
    [Field1, Field2]" shape. Not guaranteed: if the message doesn't match
    (a future API change, a different locale), this simply returns []
    and the generic guidance above still applies without the specific
    field names -- never a hard failure."""
    m = _REQUIRED_FIELD_MISSING_RE.search(message)
    if not m:
        return []
    return [f.strip() for f in m.group(1).split(",") if f.strip()]


def _is_field_ever_mapped(mapping_path, object_name, field_api_name):
    """Does object_name's mapping-doc sheet target field_api_name in ANY
    row's Target Field API cell -- regardless of that row's Migrate Data
    value? Distinct from find_unmapped_required_fields() (#49), which
    only surfaces SOURCE rows flagged Migrate Data == Yes with a still-
    blank target -- here the question is the reverse: for a TARGET field
    Salesforce just said was required and missing, was it ever chosen as
    a mapping target at all. Returns None (not True/False) if the
    mapping doc or this object's sheet doesn't exist -- "unknown," not
    "no"."""
    if not mapping_path:
        return None
    try:
        wb = openpyxl.load_workbook(mapping_path, data_only=True)
    except FileNotFoundError:
        return None
    sheet_name = _safe_sheet_name(object_name)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=4):
        target_field = row[_COL_TARGET_FIELD_API - 1].value
        if target_field and str(target_field).strip().lower() == field_api_name.strip().lower():
            return True
    return False


def _validation_rule_candidates(engine, object_name, schema="dbo"):
    """Active ValidationRule rows already on file for object_name from a
    prior analyze-org-risk run (dbo.ObjectAutomationRisk) -- [] if that
    table doesn't exist yet, or no active validation rules were found for
    this object. Read-only; never runs analyze-org-risk itself."""
    d = sql_dialect.for_engine(engine)
    if not d.table_exists(engine, schema, "ObjectAutomationRisk"):
        return []
    with engine.connect() as cx:
        rows = cx.execute(
            text(
                f"SELECT ItemName, Detail FROM {d.qualify(schema, 'ObjectAutomationRisk')} "
                "WHERE ObjectName = :obj AND CheckType = 'ValidationRule' AND IsActive = 1"
            ),
            {"obj": object_name},
        ).mappings().all()
    return [{"name": r["ItemName"], "message": r["Detail"]} for r in rows]


def triage_failures(engine, table, schema="dbo", error_column="Error",
                     object_name=None, mapping_path=None):
    """Group table's failed rows (a load table written back in place, or
    a <table>_Result table -- same calling convention as
    bulkops.build_retry_table()) by normalized error signature, and map
    each signature's error code to a likely root cause + suggested next
    step. Purely additive enrichment when object_name/mapping_path are
    given: real cross-references against dbo.ObjectAutomationRisk
    (FIELD_CUSTOM_VALIDATION_EXCEPTION) and the mapping doc's Target
    Field API column (REQUIRED_FIELD_MISSING) -- both work fine without
    those optional arguments, just with less specific guidance.

    Returns [{"signature", "code", "count", "fields", "cause",
    "next_steps", "detail"}, ...], sorted by count descending (biggest
    problem first). "detail" is a list of extra, data-driven strings
    (e.g. a matched validation rule's message, or "never mapped to any
    target field") appended past the static next_steps guidance --
    empty when no enrichment data is available."""
    d = sql_dialect.for_engine(engine)
    if not d.column_exists(engine, schema, table, error_column):
        raise ValueError(
            f"{schema}.{table} has no [{error_column}] column -- has bulkops been run against it yet?"
        )
    qualified = d.qualify(schema, table)
    with engine.connect() as cx:
        rows = cx.execute(text(
            f"SELECT {d.quote_ident(error_column)} AS err FROM {qualified} "
            f"WHERE {d.quote_ident(error_column)} IS NOT NULL"
        )).mappings().all()

    counts = {}
    for r in rows:
        signature = _normalize_error_signature(r["err"])
        counts[signature] = counts.get(signature, 0) + 1

    results = []
    for signature, count in counts.items():
        code = signature.split(":", 1)[0].strip()
        info = _ERROR_CODE_GUIDANCE.get(code)
        fields = _extract_required_fields(signature) if code == "REQUIRED_FIELD_MISSING" else []

        detail = []
        if code == "REQUIRED_FIELD_MISSING" and mapping_path and object_name:
            for field in fields:
                ever_mapped = _is_field_ever_mapped(mapping_path, object_name, field)
                if ever_mapped is True:
                    detail.append(f"{field}: already mapped in the doc -- likely a per-row source data gap, not a missing mapping.")
                elif ever_mapped is False:
                    detail.append(f"{field}: never chosen as a Target Field in the mapping doc -- likely a genuine mapping gap.")
        if code == "FIELD_CUSTOM_VALIDATION_EXCEPTION" and engine is not None and object_name:
            for rule in _validation_rule_candidates(engine, object_name, schema=schema):
                detail.append(f"Candidate rule: {rule['name']} -- \"{rule['message']}\"")

        results.append({
            "signature": signature,
            "code": code,
            "count": count,
            "fields": fields,
            "cause": info["cause"] if info else "Unrecognized error code -- no known guidance for this one yet.",
            "next_steps": info["next_steps"] if info else [
                "Read the full error text above -- this error code isn't in this tool's known-guidance list yet.",
            ],
            "detail": detail,
        })

    results.sort(key=lambda r: r["count"], reverse=True)
    return results
