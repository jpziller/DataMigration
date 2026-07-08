"""Migration run book (roadmap #16).

`dbo.BulkOpsLog` (#14) records what a *script* did automatically, but it can
never see a human turning Email Deliverability off, disabling CPQ
automation, or any other manual step -- there's no API trace of a Setup
checkbox to log. The run book is the bigger-picture document spanning both:
the full recipe for one migration project, from Pre-Migration through
Script/Transformations to Post-Migration, carried across every real pass of
it (a couple of Dev test runs, then a UAT/mock-go-live pass, then PROD).

The recipe structure itself -- section names, column headers, starter
Pre-/Post-Migration items -- lives in `docs/RUN_BOOK_TEMPLATE.md`, git-
tracked and human-editable directly, not hidden inside Python constants
(contrast `mapping_doc.py`'s `_HEADERS`). One continuous worksheet holds one
full end-to-end pass; a new pass (Dev -> UAT -> PROD) is a new *tab* in the
same workbook, created by copying the previous tab's recipe columns forward
(Item/Script name/Dependency/Critical) while blanking every execution-result
column (who, when, errors, row counts) for a fresh run.

Each section splits into recipe columns (the reusable plan, copied forward
by add_run_book_pass()) and result columns (that pass's actual execution
data, always blanked on copy) -- see _recipe_columns(). Pre-/Post-Migration
result columns will always need a human to fill them in; there's nothing to
automate there. Tying dbo.BulkOpsLog into the Script/Transformation
section's result columns automatically is the explicit next phase, not
built here.
"""
import os
import re

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "docs", "RUN_BOOK_TEMPLATE.md")
_TRANSFORMS_DIR = os.path.join(os.path.dirname(__file__), "sql", "transformations")

_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_NO_FILL = PatternFill(fill_type=None)
_HEADER_FONT = Font(bold=True)
_SECTION_TITLE_FONT = Font(bold=True, size=13)
_TOTAL_TIME_FORMAT = "[h]:mm"

_RECIPE_COLUMNS_MIGRATION = ["Item", "Critical"]
_RECIPE_COLUMNS_SCRIPT = ["Script # / Name", "Dependency"]

_SEPARATOR_CELL_RE = re.compile(r":?-+:?")


def _recipe_columns(section_name):
    if section_name.lower().startswith("script"):
        return _RECIPE_COLUMNS_SCRIPT
    return _RECIPE_COLUMNS_MIGRATION


def _is_separator_row(cells):
    return all(_SEPARATOR_CELL_RE.fullmatch(c.strip()) for c in cells)


def _parse_template(md_path):
    """Parse `## Heading` + one Markdown pipe-table per section into
    [{"name", "columns", "rows"}, ...]. rows are lists of cell strings in
    column order."""
    with open(md_path, encoding="utf-8") as fh:
        lines = fh.read().splitlines()

    sections = []
    current = None
    header_seen = False

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("## "):
            current = {"name": stripped[3:].strip(), "columns": [], "rows": []}
            sections.append(current)
            header_seen = False
            continue
        if current is None or not stripped.startswith("|"):
            continue
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        if not header_seen:
            current["columns"] = cells
            header_seen = True
            continue
        if _is_separator_row(cells):
            continue
        current["rows"].append(cells)

    return sections


def _table_exists(cx, schema, table):
    return cx.execute(text("SELECT OBJECT_ID(:t, 'U')"), {"t": f"{schema}.{table}"}).scalar() is not None


def _script_filename_for(object_name):
    if not os.path.isdir(_TRANSFORMS_DIR):
        return ""
    matches = sorted(
        f for f in os.listdir(_TRANSFORMS_DIR)
        if f.lower().endswith(".sql") and object_name.lower() in f.lower()
    )
    return matches[0] if matches else ""


def _load_order_rows(engine, object_names, schema):
    """Auto-fill the Script/Transformation section from load_order.py's
    (#2) existing dbo.ObjectLoadOrder/dbo.ObjectDependency -- same
    "prefill only what's already known, don't guess" principle as
    generate-mapping-doc's profiling auto-fill."""
    in_scope = set(object_names)

    with engine.connect() as cx:
        if not _table_exists(cx, schema, "ObjectLoadOrder"):
            raise ValueError(
                f"{schema}.ObjectLoadOrder doesn't exist yet -- run "
                f"analyze-load-order {' '.join(object_names)} first, then retry."
            )
        order_rows = cx.execute(
            text(f"SELECT ObjectName, LoadLevel, LoadSequence FROM [{schema}].[ObjectLoadOrder]")
        ).mappings().all()

        edge_rows = []
        if _table_exists(cx, schema, "ObjectDependency"):
            edge_rows = cx.execute(
                text(f"SELECT ChildObject, ParentObject FROM [{schema}].[ObjectDependency]")
            ).mappings().all()

    order_rows = [r for r in order_rows if r["ObjectName"] in in_scope]
    order_rows.sort(key=lambda r: (r["LoadSequence"] is None, r["LoadSequence"]))

    # Self-referencing fields (e.g. Account.ParentId -> Account) are their
    # own load_order.py concept (a two-pass load, not a cross-object
    # dependency) -- excluded here so they don't show up as a bogus parent
    # of themselves. Multiple lookups to the same parent object (or
    # multiple self-ref fields) are deduped with a set.
    parents_of = {}
    for e in edge_rows:
        if e["ChildObject"] == e["ParentObject"]:
            continue
        if e["ChildObject"] in in_scope and e["ParentObject"] in in_scope:
            parents_of.setdefault(e["ChildObject"], set()).add(e["ParentObject"])

    level_members = {}
    for r in order_rows:
        level_members.setdefault(r["LoadLevel"], []).append(r["ObjectName"])

    rows = []
    for r in order_rows:
        obj = r["ObjectName"]
        parents = sorted(parents_of.get(obj, []))
        siblings = sorted(n for n in level_members.get(r["LoadLevel"], []) if n != obj)

        parts = [f"After: {', '.join(parents)}" if parents else "None"]
        if siblings:
            parts.append(f"parallel with: {', '.join(siblings)}")

        rows.append({
            "Script # / Name": _script_filename_for(obj),
            "Dependency": "; ".join(parts),
        })
    return rows


def _write_section(ws, start_row, section, rows):
    """Write one section (title row, header row, data rows) starting at
    start_row. Returns the next free row (caller adds a spacer)."""
    ws.cell(row=start_row, column=1, value=section["name"]).font = _SECTION_TITLE_FONT
    header_row = start_row + 1
    columns = section["columns"]
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=col_name).font = _HEADER_FONT

    critical_idx = columns.index("Critical") + 1 if "Critical" in columns else None
    start_idx = columns.index("Start") + 1 if "Start" in columns else None
    end_idx = columns.index("End") + 1 if "End" in columns else None
    total_idx = columns.index("Total Time") + 1 if "Total Time" in columns else None

    row_idx = header_row + 1
    for row_data in rows:
        values = [row_data.get(col, "") for col in columns] if isinstance(row_data, dict) else row_data
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)

        if total_idx and start_idx and end_idx:
            start_letter, end_letter = get_column_letter(start_idx), get_column_letter(end_idx)
            cell = ws.cell(
                row=row_idx, column=total_idx,
                value=f'=IF(AND({start_letter}{row_idx}<>"",{end_letter}{row_idx}<>""),'
                      f'{end_letter}{row_idx}-{start_letter}{row_idx},"")',
            )
            cell.number_format = _TOTAL_TIME_FORMAT

        is_critical = critical_idx and str(values[critical_idx - 1] or "").strip().lower() == "yes"
        fill = _RED_FILL if is_critical else _NO_FILL
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

        row_idx += 1

    return row_idx


def _autosize_columns(ws, width_cap=60):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = min(width + 2, width_cap)


def generate_run_book(output_path, tab_name, template_path=_TEMPLATE_PATH,
                       engine=None, object_names=None, schema="dbo"):
    """Create a brand-new run-book tab from docs/RUN_BOOK_TEMPLATE.md (or a
    custom template_path). Refuses to overwrite an existing tab -- unlike
    mapping_doc.py's regenerate-in-place convention, a run-book tab holds
    live, manually-entered operational history that must never be silently
    clobbered.

    If object_names + engine are given, auto-fills the Script/Transformation
    section from analyze-load-order's (#2) results; omit them to get just
    that section's header row for manual fill-in."""
    sections = _parse_template(template_path)

    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        if tab_name in wb.sheetnames:
            raise ValueError(
                f"Tab '{tab_name}' already exists in {output_path} -- refusing to "
                "overwrite live run-book data. Pick a different --tab name."
            )
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    ws = wb.create_sheet(tab_name)

    script_rows = None
    if object_names:
        if engine is None:
            raise ValueError("object_names given without an engine")
        script_rows = _load_order_rows(engine, list(object_names), schema)

    row = 1
    for section in sections:
        rows = script_rows if (section["name"].lower().startswith("script") and script_rows is not None) else section["rows"]
        row = _write_section(ws, row, section, rows) + 1

    _autosize_columns(ws)
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    return output_path


def _find_sections_in_sheet(ws, template_sections):
    """Scan an existing tab (built by generate_run_book from this same
    template) for each section's header row + data-row range, by matching
    header cell values against the template's column lists. Relies on the
    tab having been generated from the same template -- a hand-renamed
    header means that section's rows won't be found."""
    found = []
    row = 1
    max_row = ws.max_row
    max_col = ws.max_column
    while row <= max_row:
        row_values = [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]
        matched = None
        for section in template_sections:
            cols = section["columns"]
            if row_values[:len(cols)] == cols:
                matched = section
                break
        if matched is None:
            row += 1
            continue

        # Pre-/Post-Migration share an identical column list, so the title
        # row directly above the header (always written by _write_section)
        # is the reliable way to tell which section this actually is --
        # matching on columns alone would label every Post-Migration block
        # "Pre-Migration".
        cols = matched["columns"]
        title = ws.cell(row=row - 1, column=1).value or matched["name"]
        data_start = row + 1
        data_end = data_start
        while data_end <= max_row and any(
            ws.cell(row=data_end, column=c).value not in (None, "") for c in range(1, len(cols) + 1)
        ):
            data_end += 1

        found.append({
            "name": title, "columns": cols,
            "header_row": row, "data_start": data_start, "data_end": data_end - 1,
        })
        row = data_end

    return found


def _blank_result_columns(ws, section):
    columns = section["columns"]
    recipe = set(_recipe_columns(section["name"]))
    result_idxs = [i for i, c in enumerate(columns, start=1) if c not in recipe]

    critical_idx = columns.index("Critical") + 1 if "Critical" in columns else None
    start_idx = columns.index("Start") + 1 if "Start" in columns else None
    end_idx = columns.index("End") + 1 if "End" in columns else None
    total_idx = columns.index("Total Time") + 1 if "Total Time" in columns else None

    for row_idx in range(section["data_start"], section["data_end"] + 1):
        for col_idx in result_idxs:
            # openpyxl's cell(value=None) is a no-op (indistinguishable from
            # omitting value entirely) -- must set .value directly to clear.
            ws.cell(row=row_idx, column=col_idx).value = None

        if total_idx and start_idx and end_idx:
            start_letter, end_letter = get_column_letter(start_idx), get_column_letter(end_idx)
            cell = ws.cell(
                row=row_idx, column=total_idx,
                value=f'=IF(AND({start_letter}{row_idx}<>"",{end_letter}{row_idx}<>""),'
                      f'{end_letter}{row_idx}-{start_letter}{row_idx},"")',
            )
            cell.number_format = _TOTAL_TIME_FORMAT

        critical_value = ""
        if critical_idx:
            v = ws.cell(row=row_idx, column=critical_idx).value
            critical_value = str(v).strip().lower() if v else ""
        fill = _RED_FILL if critical_value == "yes" else _NO_FILL
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def add_run_book_pass(path, from_tab, to_tab, template_path=_TEMPLATE_PATH):
    """Duplicate from_tab into a new to_tab: recipe columns (Item/Critical
    or Script # / Name/Dependency) carry forward verbatim, including any
    rows a human added by hand since generation; every result column is
    blanked for the fresh pass. Refuses to overwrite an existing to_tab."""
    wb = openpyxl.load_workbook(path)
    if from_tab not in wb.sheetnames:
        raise ValueError(f"No tab named '{from_tab}' in {path}")
    if to_tab in wb.sheetnames:
        raise ValueError(
            f"Tab '{to_tab}' already exists in {path} -- refusing to overwrite live run-book data."
        )

    dst = wb.copy_worksheet(wb[from_tab])
    dst.title = to_tab

    template_sections = _parse_template(template_path)
    for section in _find_sections_in_sheet(dst, template_sections):
        _blank_result_columns(dst, section)

    wb.save(path)
    return path
