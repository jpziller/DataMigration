"""Field-mapping spreadsheet tool (roadmap #3).

Column structure is modeled on a real-world field-inventory-and-mapping
template (structure/column-names only -- reviewed for format, not content):
one row per SOURCE field, a block of migration-decision columns (populated
%, migrate data/field Y-N, business review/decision), then a blank Target
block (Object/Field API/Label/Type/Description/Notes) for a human to fill
in once a mapping is actually decided. This is the opposite orientation
from an earlier version of this tool (which listed one row per TARGET
field) -- row-per-source-field is the right shape, since the real workflow
is "for each source field, decide if/how it maps," not the reverse.

generate_mapping_workbook() builds this from a SQL Server source table's
real columns (INFORMATION_SCHEMA) against a named Salesforce target object.
If profiling data already exists for that table (see profiling.py),
"Data Profile Populated On"/"Data Profile %" are pre-filled from it -- those
columns exist for exactly that purpose. This does NOT auto-guess the
mapping itself (that's a separate roadmap item) -- only the structure.

check_mapping_balance() diffs a filled-in mapping doc against the actual
`sql/transformations/*.sql` load-table-building code in both directions, so
the human-readable doc and the executable transform can't silently drift
apart:
  - documented_not_implemented: the doc shows a target field as mapped
    (Target block's Field API is filled in), but the transform's INSERT
    INTO column list doesn't populate it.
  - implemented_not_documented: the transform populates a column with no
    row showing it as mapped. This also catches a transform populating a
    field that doesn't exist in the object's current describe() at all --
    a nonexistent field can't have a row, so it always surfaces here.
"""
import os
import re

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import text

import git_info as gi
import script_numbering as sn
import sql_dialect

_HYPERLINK_FONT = Font(color="0563C1", underline="single")

_INSERT_INTO_RE = re.compile(
    r'INSERT\s+INTO\s+(?:(?:\[\w+\]|"\w+"|\w+)\.)?'
    r'(?:\[(?P<table_br>\w+)\]|"(?P<table_dq>\w+)"|(?P<table_bare>\w+))'
    r"\s*\((?P<cols>[^)]+)\)",
    re.IGNORECASE,
)
# sql_dialect.py's own create_table_as_select_sql() -- the actual pattern
# this project's real transform scripts use to build a *_Load table, on
# either backend -- is neither of these forms; a plain INSERT INTO (...)
# VALUES-shaped statement (what _INSERT_INTO_RE above matches) never
# appears in a real sql/transformations/*.sql script at all. Found via a
# real dogfood run: check-mapping-balance/assess-migration-readiness's
# mapping_balance gate raised "No INSERT INTO statement found" against
# every one of this project's own real, working scripts.
#
# mssql: SELECT <cols> INTO [schema].[table] FROM ...
_SELECT_INTO_RE = re.compile(
    r'SELECT\s+(?P<cols>.*?)\s+INTO\s+(?:(?:\[\w+\]|"\w+"|\w+)\.)?'
    r'(?:\[(?P<table_br>\w+)\]|"(?P<table_dq>\w+)"|(?P<table_bare>\w+))',
    re.IGNORECASE | re.DOTALL,
)
# sqlite: CREATE TABLE [schema].[table] AS SELECT <cols> FROM ...
_CREATE_TABLE_AS_SELECT_RE = re.compile(
    r'CREATE\s+TABLE\s+(?:(?:\[\w+\]|"\w+"|\w+)\.)?'
    r'(?:\[(?P<table_br>\w+)\]|"(?P<table_dq>\w+)"|(?P<table_bare>\w+))'
    r'\s+AS\s+SELECT\s+(?P<cols>.*?)\s+FROM\b',
    re.IGNORECASE | re.DOTALL,
)
_AS_ALIAS_RE = re.compile(r'\bAS\s+(\[[^\]]+\]|"[^"]+"|\w+)\s*$', re.IGNORECASE)
_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")

_HEADERS = [
    "Source Object", "Field API", "Field Label", "Data Type", "Description",
    "Data Profile Populated On", "Data Profile %", "Notes",
    "Migrate Data", "Migrate Field", "Biz Review Req", "Biz Decision",
    None,  # spacer column between source and target blocks
    "Target Object", "Field API", "Field Label", "Data Type", "Description", "Notes",
]
_TARGET_FIELD_API_COL = 15  # 1-indexed position of the Target block's "Field API" column

# Named column positions for apply_auto_map_suggestions() -- see _HEADERS above.
_COL_SOURCE_FIELD_API = 2
_COL_SOURCE_NOTES = 8
_COL_MIGRATE_DATA = 9
_COL_TARGET_OBJECT = 14
_COL_TARGET_FIELD_API = 15
_COL_TARGET_FIELD_LABEL = 16
_COL_TARGET_DATA_TYPE = 17


def _safe_sheet_name(name):
    return _INVALID_SHEET_CHARS.sub("_", name)[:31]


def generate_mapping_workbook(sf, target_object, output_path, engine, source_table, schema="dbo"):
    d = sql_dialect.for_engine(engine)
    source_cols = [{"COLUMN_NAME": name, "DATA_TYPE": dtype}
                   for name, dtype in d.list_columns(engine, schema, source_table)]

    if not source_cols:
        raise ValueError(f"No such table: {schema}.{source_table}")

    profile_by_field = {}
    if d.table_exists(engine, schema, "FieldProfile"):
        qualified = d.qualify(schema, "FieldProfile")
        with engine.connect() as cx:
            profile_rows = cx.execute(
                text(
                    f"SELECT FieldName, TotalRows, PopulatedCount, PopulatedPct "
                    f"FROM {qualified} "
                    "WHERE ObjectOrTable = :table AND SourceType = 'sql_table'"
                ),
                {"table": source_table},
            ).mappings().all()
        profile_by_field = {r["FieldName"]: r for r in profile_rows}

    sheet_name = _safe_sheet_name(target_object)
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value="Source Object:")
    ws.cell(row=1, column=2, value=source_table)
    ws.cell(row=1, column=3, value="Target Object:")
    ws.cell(row=1, column=4, value=target_object)
    ws.cell(row=1, column=5, value="Transform Script:")
    # Deliberately left blank here -- the transform doesn't exist yet at
    # generate-mapping-doc time in the standard workflow (mapping comes
    # before building the transform). set_transform_script() fills this in
    # later, once the real script has actually been built.

    for col_idx, header in enumerate(_HEADERS, start=1):
        if header is not None:
            ws.cell(row=3, column=col_idx, value=header)

    row_idx = 4
    for col in source_cols:
        name = col["COLUMN_NAME"]
        profile = profile_by_field.get(name)
        ws.cell(row=row_idx, column=1, value=source_table)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=4, value=col["DATA_TYPE"])
        if profile:
            ws.cell(row=row_idx, column=6, value=f"{profile['PopulatedCount']} of {profile['TotalRows']}")
            pct = profile["PopulatedPct"]
            ws.cell(row=row_idx, column=7, value=float(pct) if pct is not None else None)
        row_idx += 1

    wb.save(output_path)
    return output_path


def set_transform_script(mapping_path, target_object, script_subdir="transformations", repo_root=None):
    """Fill in the "Transform Script:" header field (row 1) on
    target_object's sheet with the real script that implements it --
    resolved from sql/<script_subdir>/ the same way migration_run_book.py's
    Load phase does (highest-numbered match wins), never hand-typed.
    script_subdir is "transformations" (default) or "source_ingestion",
    matching next-script-number's own --dir choices. repo_root defaults to
    this repo (exposed only so tests can point it at a tmp_path instead of
    the real sql/ tree). A real hyperlink to that file at the current
    commit is attached too, when this repo has a GitHub remote -- same
    breadcrumb convention the Migration Run Book uses.

    Deliberately a separate step, not part of generate_mapping_workbook():
    in the standard workflow, mapping comes before the transform is built,
    so the script genuinely doesn't exist yet at that point. Raises if no
    matching script is found -- this step only makes sense to run after
    the real script exists, not as a guess at what it will be named.
    Prints a warning (doesn't raise) when MORE THAN ONE real, distinct
    script matches target_object -- found in review, then found to be
    more nuanced than a simple bug once an existing test was checked:
    "two matches" covers both a genuinely harmless case (an old,
    superseded draft alongside the current script -- e.g. this file's
    own test_set_transform_script_prefers_highest_numbered_match fixture,
    010_account_load.sql superseded by 040_account_load.sql, where
    highest-number-wins is exactly the wanted behavior) and a genuinely
    dangerous one (two DIFFERENT, both-still-real scripts for the same
    object via different routing branches -- this project's own
    GiftCommitment, split into a Recurring-Donation-routed script and a
    separate Opportunity-routed one). Nothing in the filenames/numbers
    alone reliably distinguishes the two cases, so raising unconditionally
    would break the first, legitimate pattern -- a warning surfaces the
    second, dangerous pattern for a human to notice without breaking the
    first. Still returns the highest-numbered match either way, same
    resolution as before this review."""
    wb = openpyxl.load_workbook(mapping_path)
    sheet_name = _safe_sheet_name(target_object)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No sheet named '{sheet_name}' in {mapping_path}")
    ws = wb[sheet_name]

    repo_root = repo_root if repo_root is not None else os.path.dirname(__file__)
    directory = os.path.join(repo_root, "sql", script_subdir)
    # wb.sheetnames -- every object this project has already built a
    # mapping doc for -- doubles as a known_objects set so a compound-name
    # script can't silently outrank the real script for a shorter object
    # it happens to embed (ROADMAP #76). Not a full project registry (an
    # object never mapped in this workbook won't be in it), but a real,
    # already-available improvement over no known_objects at all.
    candidates = sn.script_candidates_for(target_object, directory, known_objects=wb.sheetnames)
    if not candidates:
        raise ValueError(
            f"No transform script for '{target_object}' found in sql/{script_subdir}/ -- "
            "build the transform first, then run this."
        )
    if len(candidates) > 1:
        print(
            f"Warning: more than one transform script matches '{target_object}' in "
            f"sql/{script_subdir}/: {candidates} -- using the highest-numbered "
            f"({candidates[-1]}). If these are genuinely different scripts (e.g. "
            "different source-routing branches for the same target object, not one "
            "superseding another), confirm this is really the one this mapping doc "
            "is for before trusting the link written below."
        )
    filename = candidates[-1]

    ws.cell(row=1, column=5, value="Transform Script:")
    cell = ws.cell(row=1, column=6, value=filename)
    info = gi.get_git_info()
    url = gi.github_url(info["remote_url"]) if info else None
    if url:
        cell.hyperlink = f'{url}/blob/{info["commit_sha"]}/sql/{script_subdir}/{filename}'
        cell.font = _HYPERLINK_FONT

    wb.save(mapping_path)
    return filename


def apply_auto_map_suggestions(mapping_path, object_name, target_object, suggestions):
    """Write auto_mapper.py's suggestions into an existing mapping doc's
    Target block + source Notes + Migrate Data columns.

    Never overwrites a row where a human has already filled in the Target
    block's Field API -- a human decision always wins over a suggestion,
    silently or otherwise.
    """
    wb = openpyxl.load_workbook(mapping_path)
    sheet_name = _safe_sheet_name(object_name)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No sheet named '{sheet_name}' in {mapping_path}")
    ws = wb[sheet_name]

    suggestions_by_field = {s["source_field"]: s for s in suggestions}
    applied, skipped_human = 0, 0

    for row in ws.iter_rows(min_row=4):
        source_field = row[_COL_SOURCE_FIELD_API - 1].value
        if source_field not in suggestions_by_field:
            continue

        existing_target = row[_COL_TARGET_FIELD_API - 1].value
        if existing_target is not None and str(existing_target).strip():
            skipped_human += 1
            continue

        s = suggestions_by_field[source_field]
        row_idx = row[0].row
        if s["target_field"]:
            ws.cell(row=row_idx, column=_COL_TARGET_OBJECT, value=target_object)
            ws.cell(row=row_idx, column=_COL_TARGET_FIELD_API, value=s["target_field"])
            ws.cell(row=row_idx, column=_COL_TARGET_FIELD_LABEL, value=s["target_label"])
            ws.cell(row=row_idx, column=_COL_TARGET_DATA_TYPE, value=s["target_type"])
        ws.cell(row=row_idx, column=_COL_SOURCE_NOTES, value=s["rationale"])
        ws.cell(row=row_idx, column=_COL_MIGRATE_DATA, value=s["migrate_recommended"])
        applied += 1

    wb.save(mapping_path)
    return {"applied": applied, "skipped_human_filled": skipped_human}


def find_unmapped_required_fields(mapping_path, object_name):
    """Every row in object_name's sheet flagged Migrate Data == 'Yes' with
    no Target Field API ever chosen (roadmap #49) -- a silent gap that
    would otherwise only surface once the transform is built and a field
    is simply missing. Read-only; refines #3/#10.

    Returns [{"source_field", "notes"}, ...], in row order."""
    wb = openpyxl.load_workbook(mapping_path, data_only=True)
    sheet_name = _safe_sheet_name(object_name)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No sheet named '{sheet_name}' in {mapping_path}")
    ws = wb[sheet_name]

    gaps = []
    for row in ws.iter_rows(min_row=4):
        migrate = row[_COL_MIGRATE_DATA - 1].value
        if migrate is None or str(migrate).strip() != "Yes":
            continue
        target_field = row[_COL_TARGET_FIELD_API - 1].value
        if target_field is not None and str(target_field).strip():
            continue
        source_field = row[_COL_SOURCE_FIELD_API - 1].value
        if not source_field:
            continue
        gaps.append({
            "source_field": str(source_field).strip(),
            "notes": row[_COL_SOURCE_NOTES - 1].value,
        })
    return gaps


_BLOCK_COMMENT_RE = re.compile(r"/\*.*?\*/", re.DOTALL)
_LINE_COMMENT_RE = re.compile(r"--[^\n]*")


def _strip_sql_comments(sql_text):
    """Strip /* ... */ block comments and -- line comments before any
    pattern match below -- found via a real dogfood run, a genuinely
    embarrassing bug: this project's own header comments describe the
    SELECT...INTO port in English prose (literally "SELECT ... INTO is
    the equivalent"), and _SELECT_INTO_RE matched that PROSE as if it
    were real SQL (extracting "is" as the table name and comment text as
    the column list) before this stripping step existed. A header comment
    mentioning any of these keywords in passing must never be mistaken
    for the real statement the same file's actual code contains."""
    return _LINE_COMMENT_RE.sub("", _BLOCK_COMMENT_RE.sub("", sql_text))


def _split_top_level_commas(text):
    """Split a SELECT column list on commas that aren't nested inside
    parentheses -- a plain text.split(",") would wrongly split a single
    column expression like CAST(x AS NVARCHAR(50)) or CONVERT(x, y) into
    multiple pieces."""
    parts, depth, current = [], 0, []
    for ch in text:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "," and depth == 0:
            parts.append("".join(current))
            current = []
        else:
            current.append(ch)
    if current:
        parts.append("".join(current))
    return parts


def _column_name_from_select_expr(expr):
    """The implemented column name for one SELECT column expression: the
    alias after AS if present (e.g. "_MockRowId AS LoadId" -> "LoadId",
    "CAST(x AS NVARCHAR(50)) AS MigrationID__c" -> "MigrationID__c" --
    _AS_ALIAS_RE anchors to the END of the expression so an AS inside a
    CAST(...)'s own type-conversion syntax is never mistaken for the
    column's real alias), otherwise the bare/qualified column name itself
    (e.g. "m.Subject" -> "Subject", "Name" -> "Name")."""
    expr = expr.strip()
    m = _AS_ALIAS_RE.search(expr)
    if m:
        return m.group(1).strip('[]"')
    return expr.split(".")[-1].strip().strip('[]"')


def _extract_select_style_columns(sql_text, table_name, pattern):
    for match in pattern.finditer(sql_text):
        found_table = (
            match.group("table_br") or match.group("table_dq") or match.group("table_bare")
        )
        if table_name is not None and found_table.lower() != table_name.lower():
            continue
        col_list = match.group("cols")
        return [_column_name_from_select_expr(c) for c in _split_top_level_commas(col_list)]
    return None


def extract_insert_columns(sql_text, table_name=None):
    """Return the implemented column list for a *_Load table build in the
    given SQL text -- the first statement matching table_name
    (case-insensitive, schema prefix/brackets/double-quotes ignored), or
    the first match found if table_name is None. Recognizes all three real
    patterns this project's own transform scripts use:
      - INSERT INTO table (col1, col2, ...) -- a plain parenthesized list.
      - SELECT col1, col2 AS alias, ... INTO table FROM ... -- mssql's own
        canonical *_Load-building idiom (sql_dialect.py's
        MssqlDialect.create_table_as_select_sql()); found via a real
        dogfood run that this form was never recognized at all before,
        so check-mapping-balance raised "No INSERT INTO statement found"
        against every one of this project's own real, working scripts.
      - CREATE TABLE table AS SELECT col1, col2 AS alias, ... FROM ... --
        the equivalent sqlite idiom.
    Recognizes SQL Server's [bracket] quoting and SQLite/ANSI's
    "double quote" quoting (SqliteDialect.qualify()'s own output shape),
    not just bracket-or-bare. Returns None if no match is found in any
    of the three forms."""
    sql_text = _strip_sql_comments(sql_text)
    for match in _INSERT_INTO_RE.finditer(sql_text):
        found_table = (
            match.group("table_br") or match.group("table_dq") or match.group("table_bare")
        )
        col_list = match.group("cols")
        if table_name is None or found_table.lower() == table_name.lower():
            return [c.strip().strip('[]"') for c in col_list.split(",")]

    result = _extract_select_style_columns(sql_text, table_name, _SELECT_INTO_RE)
    if result is not None:
        return result
    return _extract_select_style_columns(sql_text, table_name, _CREATE_TABLE_AS_SELECT_RE)


def check_mapping_balance(sf, mapping_path, object_name, transform_sql_path, load_table_name=None):
    wb = openpyxl.load_workbook(mapping_path, data_only=True)
    sheet_name = _safe_sheet_name(object_name)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No sheet named '{sheet_name}' in {mapping_path}")
    ws = wb[sheet_name]

    # Hard rule 14: a target field chosen by more than one row within this
    # ONE sheet is a real problem (an ambiguous INSERT down the line) --
    # different sheets/scripts targeting the same field is fine and
    # expected (e.g. two source systems feeding the same object), so this
    # is deliberately scoped to one sheet, not the whole workbook. Capture
    # duplicates before collapsing to a set for the balance check below.
    source_fields_by_target = {}
    for row in ws.iter_rows(min_row=4):
        target = row[_TARGET_FIELD_API_COL - 1].value
        if target is None or not str(target).strip():
            continue
        target = str(target).strip()
        source = row[_COL_SOURCE_FIELD_API - 1].value
        source_fields_by_target.setdefault(target, []).append(str(source).strip() if source else None)
    duplicate_target_fields = {t: srcs for t, srcs in source_fields_by_target.items() if len(srcs) > 1}
    documented = set(source_fields_by_target)

    with open(transform_sql_path, encoding="utf-8") as fh:
        sql_text = fh.read()

    implemented_cols = extract_insert_columns(sql_text, load_table_name)
    if implemented_cols is None:
        raise ValueError(f"No INSERT INTO statement found in {transform_sql_path}")
    # Same rule 14 check, transform side: a column named more than once in
    # ONE script's own INSERT INTO/CREATE TABLE column list would break
    # the actual SQL outright -- catch it before collapsing to a set.
    duplicate_implemented_columns = sorted(
        {c for c in implemented_cols if implemented_cols.count(c) > 1}
    )
    implemented = set(implemented_cols)

    # Unlike the mapping doc (free text) and the transform's INSERT list
    # (whatever the SQL author typed), describe() is ground truth for what
    # fields actually exist on the target object -- cross-check both sets
    # against it so a typo'd/removed/never-deployed field name gets flagged
    # explicitly, rather than only showing up as an ordinary imbalance.
    real_fields = {f["name"] for f in getattr(sf, object_name).describe()["fields"]}
    not_real_field = (documented | implemented) - real_fields

    # A field that isn't real at all is the more urgent finding -- report it
    # there only, not also as an ordinary documented/implemented imbalance.
    documented -= not_real_field
    implemented -= not_real_field

    return {
        "documented_not_implemented": sorted(documented - implemented),
        "implemented_not_documented": sorted(implemented - documented),
        "not_a_real_field": sorted(not_real_field),
        "duplicate_target_fields": duplicate_target_fields,
        "duplicate_implemented_columns": duplicate_implemented_columns,
    }
