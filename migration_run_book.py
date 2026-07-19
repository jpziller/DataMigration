"""Migration Run Book (roadmap #16).

`dbo.BulkOpsLog` (#14) records what a *script* did automatically, but it can
never see a human turning Email Deliverability off, disabling CPQ
automation, or any other manual step -- there's no API trace of a Setup
checkbox to log. The Migration Run Book is the bigger-picture document
spanning both: the full recipe for one migration project, from
Pre-Migration through Load steps to Post-Migration, carried across every
real pass of it (a couple of Dev test runs, then a UAT/mock-go-live pass,
then PROD).

Structure is a direct mirror (structure only, reviewed for column names/
layout, never content) of a real client's in-production migration-status
tracking tab: **one unified table** with a single fixed column schema
(_COLUMNS) used for every phase, rather than a different column set per
section. Phases (Pre-Migration, Load, Post-Migration, or a project's own
custom ones) are marked by a single banner row -- one full-width, dark
navy/white-font row with the phase name in its first cell -- not a
repeated header. The one true column-header row is written once, right
after this module's own breadcrumb block (see below).

Recipe columns (_RECIPE_COLUMNS: Stage/Object/Dependency/Critical/JIRA
Ticket Link) are the reusable plan, carried forward verbatim by
add_migration_run_book_pass() into a new pass's tab. Every other column is
a result -- that pass's actual execution data -- and gets blanked (Status
reset to "Not Started", not left empty) for the fresh run. Pre-/Post-
Migration result columns will always need a human to fill them in --
there's nothing to automate there. Tying dbo.BulkOpsLog into the Load
phase's result columns automatically is the explicit next phase, not
built here.

Status is a real Excel dropdown (_STATUS_VALUES) driven by genuine
conditional-formatting rules keyed to the cell's value -- colors update
live if a human changes the dropdown later, not a one-time paint. Same
mechanism for Critical. Both are over-provisioned to _MAX_DATA_ROW so rows
a human adds later still pick up the formatting, matching the real file's
own practice.

Every tab also gets a fixed-height breadcrumb block (rows 1-7, see
_HEADER_ROW_*) above the table: Project, Source/Target Environment, a
Git Repository link, the exact commit/branch this pass's scripts came
from, a link to those scripts at that commit, and (optionally) a link to
the project's ticket system (JIRA or whatever a given engagement actually
uses -- a per-row "JIRA Ticket Link" column also exists, for a specific
story/bug tied to one step, distinct from this project-level link).
Unlike the phase templates, this breadcrumb block is a Python-defined
layout, not sourced from docs/MIGRATION_RUN_BOOK_TEMPLATE.md -- every
value is either computed (Git info) or supplied at generation time. The
height is fixed (blank cell, not a skipped row, when a value is unknown)
specifically so add_migration_run_book_pass() can refresh header values in
place without shifting the table already copied below it.
"""
import os
import re

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import text

import git_info as gi
import script_numbering as sn
import sql_dialect

_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "docs", "MIGRATION_RUN_BOOK_TEMPLATE.md")
_TRANSFORMS_DIR = os.path.join(os.path.dirname(__file__), "sql", "transformations")

# The one shared column schema every phase uses -- mirrors a real client
# file's actual header, plus this framework's own additions. "Object" is
# kept as the literal name (not "Item"/"Script Name") to match that file
# exactly, even though it really means "step/task/script name", not a
# Salesforce object API name.
_COLUMNS = [
    "Stage", "Object", "Dependency", "Status", "Critical",
    "Person Responsible", "Begin Time", "End Time", "Execution Time",
    "JIRA Ticket Link", "Notes", "Total Records", "Success Records",
    "Failed Records", "Success Percent", "Error Details",
]
# Dependency, Critical, and JIRA Ticket Link are this framework's own
# additions (the real file has neither) -- kept here as recipe columns
# since a step's dependency, ahead-of-time risk flag, and associated
# ticket don't change from one pass to the next, unlike its execution
# result.
_RECIPE_COLUMNS = ["Stage", "Object", "Dependency", "Critical", "JIRA Ticket Link"]

_STATUS_VALUES = ["Not Started", "N/A", "In Process", "Completed", "Issue"]
_STATUS_FILLS = {
    "In Process": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),
    "N/A": PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
    "Completed": PatternFill(start_color="FF375623", end_color="FF375623", fill_type="solid"),
    "Issue": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
}
_STATUS_FONTS = {
    "Completed": Font(color="FFFFFFFF"),
    "Issue": Font(color="FFFFFFFF"),
}
# Distinct, more saturated red than Critical's fill so the two flags (an
# ahead-of-time risk marker vs. a live Status value) stay visually
# distinguishable from each other.
_CRITICAL_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

_BANNER_FILL = PatternFill(start_color="FF0D2C39", end_color="FF0D2C39", fill_type="solid")
_BANNER_FONT = Font(color="FFFFFFFF")
_TABLE_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_HEADER_FONT = Font(bold=True)
_HYPERLINK_FONT = Font(color="0563C1", underline="single")

# Conditional formatting/data validation ranges are over-provisioned this
# far past the last written row, matching the real file's own practice
# (its ranges ran to row 1062), so rows a human adds later still pick up
# the Status dropdown and live coloring.
_MAX_DATA_ROW = 1000

_SEPARATOR_CELL_RE = re.compile(r":?-+:?")

# Fixed breadcrumb row positions (rows 1-7) -- see module docstring for
# why this is a constant height rather than "skip if unknown".
_HEADER_ROW_PROJECT = 1
_HEADER_ROW_SOURCE_ENV = 2
_HEADER_ROW_TARGET_ENV = 3
_HEADER_ROW_GIT_REPO = 4
_HEADER_ROW_COMMIT = 5
_HEADER_ROW_SCRIPTS_LINK = 6
_HEADER_ROW_TICKET = 7
_HEADER_ROW_LAST_SYNCED_LOG_ID = 8
_HEADER_ROW_LAST_SYNCED_SOURCE_LOG_ID = 9   # previously a blank spacer row --
# repurposed for SourceIngestionLog's own independent watermark (roadmap #46)
# rather than shifting _TABLE_HEADER_ROW/_FIRST_DATA_ROW, which would
# desync any already-generated workbook's existing row layout.
_TABLE_HEADER_ROW = 10
_FIRST_DATA_ROW = 11


def _is_separator_row(cells):
    return all(_SEPARATOR_CELL_RE.fullmatch(c.strip()) for c in cells)


def _parse_template(md_path):
    """Parse `## Heading` (a phase/banner) + one Markdown pipe-table per
    heading -- all sharing the _COLUMNS schema -- into
    [{"name", "rows"}, ...]. rows are lists of cell strings in _COLUMNS
    order. The MD table's own header row is validated against _COLUMNS
    (not otherwise used) so a hand-edited template can't silently drift
    from what the code actually writes."""
    with open(md_path, encoding="utf-8") as fh:
        lines = fh.read().splitlines()

    phases = []
    current = None
    header_seen = False

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("## "):
            current = {"name": stripped[3:].strip(), "rows": []}
            phases.append(current)
            header_seen = False
            continue
        if current is None or not stripped.startswith("|"):
            continue
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        if not header_seen:
            if cells != _COLUMNS:
                raise ValueError(
                    f"{md_path} phase '{current['name']}' table header {cells} "
                    f"doesn't match the expected schema {_COLUMNS}"
                )
            header_seen = True
            continue
        if _is_separator_row(cells):
            continue
        # Same drift protection as the header check above -- a data row
        # with the wrong cell count would otherwise be silently truncated/
        # misaligned by _write_data_row's zip() (found live: an extra
        # trailing empty cell on every starter row went unnoticed).
        if len(cells) != len(_COLUMNS):
            raise ValueError(
                f"{md_path} phase '{current['name']}' has a data row with "
                f"{len(cells)} cell(s), expected {len(_COLUMNS)}: {cells[:3]}..."
            )
        current["rows"].append(cells)

    return phases


def _save_workbook(wb, path):
    """Save with a clear message when the file is locked -- on Windows an
    open-in-Excel workbook raises a bare PermissionError, and 'close the
    file' is a much better answer than a traceback (a run book is exactly
    the kind of file someone has open while working)."""
    try:
        wb.save(path)
    except PermissionError as e:
        raise ValueError(
            f"Can't write {path} -- it's locked, most likely open in Excel. "
            "Close it there and re-run."
        ) from e


def _write_breadcrumb_header(ws, git_info, project_name=None, source_env=None, target_env=None,
                              ticket_url=None, ticket_label="JIRA"):
    """Write the fixed-height breadcrumb block (rows 1-8). "Last Synced Log
    Id" (row 8) is always written blank here -- it's per-pass state
    managed exclusively by sync_run_book_from_log(), not something this
    function or its caller ever supplies, and (like Target Environment)
    it deliberately never carries forward: a fresh tab or a new pass
    hasn't had any of its own runs logged yet."""
    github_url = gi.github_url(git_info["remote_url"]) if git_info else None
    commit_label = f'{git_info["commit_sha"][:8]} ({git_info["branch"]})' if git_info else None
    scripts_url = (
        f'{github_url}/tree/{git_info["commit_sha"]}/sql/transformations'
        if (git_info and github_url) else None
    )

    rows = [
        (_HEADER_ROW_PROJECT, "Project", project_name, None),
        (_HEADER_ROW_SOURCE_ENV, "Source Environment", source_env, None),
        (_HEADER_ROW_TARGET_ENV, "Target Environment", target_env, None),
        (_HEADER_ROW_GIT_REPO, "Git Repository", github_url, github_url),
        (_HEADER_ROW_COMMIT, "Commit / Branch", commit_label, None),
        (_HEADER_ROW_SCRIPTS_LINK, "Scripts (as of this commit)", scripts_url, scripts_url),
        (_HEADER_ROW_TICKET, f"{ticket_label} Project", ticket_url, ticket_url),
        (_HEADER_ROW_LAST_SYNCED_LOG_ID, "Last Synced Log Id", None, None),
        (_HEADER_ROW_LAST_SYNCED_SOURCE_LOG_ID, "Last Synced Source Ingestion Log Id", None, None),
    ]
    for row_idx, label, value, hyperlink in rows:
        ws.cell(row=row_idx, column=1, value=label).font = _HEADER_FONT
        # openpyxl's cell(value=None) is a no-op (indistinguishable from
        # omitting value entirely) -- setting .value directly is required
        # to actually clear an already-populated cell (matters here since
        # add_migration_run_book_pass calls this on a tab copy_worksheet()
        # already populated from the source tab).
        cell = ws.cell(row=row_idx, column=2)
        cell.value = value
        if hyperlink:
            cell.hyperlink = hyperlink
            cell.font = _HYPERLINK_FONT


def _read_breadcrumb_header(ws):
    """Read back the carry-forward-able breadcrumb values from an existing
    tab -- Target Environment and Last Synced Log Id are deliberately
    excluded (see add_migration_run_book_pass: Dev/UAT/PROD are different
    orgs and each pass's own sync history starts fresh, neither is ever
    silently reused)."""
    ticket_cell_label = ws.cell(row=_HEADER_ROW_TICKET, column=1).value or ""
    ticket_label = (
        ticket_cell_label[:-len(" Project")] if ticket_cell_label.endswith(" Project") else "JIRA"
    )
    return {
        "project_name": ws.cell(row=_HEADER_ROW_PROJECT, column=2).value,
        "source_env": ws.cell(row=_HEADER_ROW_SOURCE_ENV, column=2).value,
        "ticket_url": ws.cell(row=_HEADER_ROW_TICKET, column=2).value,
        "ticket_label": ticket_label,
    }


def _write_table_header(ws, row):
    for col_idx, col_name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = _TABLE_HEADER_FONT
        cell.fill = _BANNER_FILL


def _write_banner(ws, row, label):
    ws.cell(row=row, column=1, value=label)
    for col_idx in range(1, len(_COLUMNS) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = _BANNER_FILL
        cell.font = _BANNER_FONT


def _write_data_row(ws, row_idx, row_data):
    """row_data: dict keyed by column name (missing keys -> blank) or a
    plain list already in _COLUMNS order. A "__hyperlink__" key applies
    that URL to the Object cell. Live Status/Critical coloring comes from
    the conditional-formatting rules applied once over the whole data
    range (_apply_conditional_formatting), not painted per-cell here."""
    values = [row_data.get(col, "") for col in _COLUMNS] if isinstance(row_data, dict) else row_data
    hyperlink_url = row_data.get("__hyperlink__") if isinstance(row_data, dict) else None
    for col_idx, (col_name, value) in enumerate(zip(_COLUMNS, values), start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)
        if hyperlink_url and col_name == "Object":
            cell.hyperlink = hyperlink_url
            cell.font = _HYPERLINK_FONT


def _apply_conditional_formatting(ws):
    """Real Excel conditional-formatting rules for Status/Critical (colors
    update live if a human changes the value later -- not a one-time
    paint), plus the Status dropdown. No clearing step is needed before
    re-applying on a copied tab: confirmed against the installed openpyxl
    (3.1.5), copy_worksheet() does not carry conditional formatting or
    data validations to the copy at all -- which is exactly why
    add_migration_run_book_pass() must call this on every new pass. (An
    earlier version "cleared" via ws._data_validations = [], an attribute
    that doesn't exist -- a silent no-op, removed rather than kept as
    false reassurance.)"""
    status_col = _COLUMNS.index("Status") + 1
    critical_col = _COLUMNS.index("Critical") + 1
    status_letter = get_column_letter(status_col)
    critical_letter = get_column_letter(critical_col)
    status_range = f"{status_letter}{_FIRST_DATA_ROW}:{status_letter}{_MAX_DATA_ROW}"
    critical_range = f"{critical_letter}{_FIRST_DATA_ROW}:{critical_letter}{_MAX_DATA_ROW}"

    for value, fill in _STATUS_FILLS.items():
        ws.conditional_formatting.add(
            status_range,
            CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill, font=_STATUS_FONTS.get(value)),
        )
    ws.conditional_formatting.add(
        critical_range,
        CellIsRule(operator="equal", formula=['"Yes"'], fill=_CRITICAL_FILL),
    )

    dv = DataValidation(type="list", formula1=f'"{",".join(_STATUS_VALUES)}"', allow_blank=True)
    dv.add(status_range)
    ws.add_data_validation(dv)


def _size_columns(ws):
    """Fixed, deliberate widths mirroring the real file's proportions
    (Object/Notes hold long free text and were ~106/~127 chars wide there)
    rather than pure content-based autosize, so a fresh tab still reads
    well before anyone's typed anything into it."""
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))

    min_widths = {"Object": 60, "Notes": 70}
    for idx, col_name in enumerate(_COLUMNS, start=1):
        letter = get_column_letter(idx)
        width = max(widths.get(letter, 0) + 2, min_widths.get(col_name, 14))
        ws.column_dimensions[letter].width = min(width, 90)


def _load_order_rows(engine, object_names, schema, git_info=None):
    """Auto-fill Load-phase rows from load_order.py's (#2) existing
    dbo.ObjectLoadOrder/dbo.ObjectDependency -- same "prefill only what's
    already known, don't guess" principle as generate-mapping-doc's
    profiling auto-fill. When git_info resolves to a GitHub remote, a
    matched script filename also gets a real hyperlink to that exact file
    at the pinned commit."""
    in_scope = set(object_names)
    github_url = gi.github_url(git_info["remote_url"]) if git_info else None

    d = sql_dialect.for_engine(engine)
    if not d.table_exists(engine, schema, "ObjectLoadOrder"):
        raise ValueError(
            f"{schema}.ObjectLoadOrder doesn't exist yet -- run "
            f"analyze-load-order {' '.join(object_names)} first, then retry."
        )
    with engine.connect() as cx:
        order_rows = cx.execute(
            text(f"SELECT ObjectName, LoadLevel, LoadSequence FROM {d.qualify(schema, 'ObjectLoadOrder')}")
        ).mappings().all()

        edge_rows = []
        if d.table_exists(engine, schema, "ObjectDependency"):
            edge_rows = cx.execute(
                text(f"SELECT ChildObject, ParentObject FROM {d.qualify(schema, 'ObjectDependency')}")
            ).mappings().all()

    # Normalized to lowercase keys once here rather than accessed by exact
    # case throughout this function -- found via live Postgres testing
    # (see orchestrator.py's own _row_to_current() fix for the identical
    # issue): an unquoted column comes back lowercased in a Postgres
    # result set even though SQL Server/SQLite both preserve the
    # originally-declared case.
    order_rows = [sql_dialect.lower_keys(r) for r in order_rows]
    edge_rows = [sql_dialect.lower_keys(r) for r in edge_rows]

    order_rows = [r for r in order_rows if r["objectname"] in in_scope]
    order_rows.sort(key=lambda r: (r["loadsequence"] is None, r["loadsequence"]))

    # Self-referencing fields (e.g. Account.ParentId -> Account) are their
    # own load_order.py concept (a two-pass load, not a cross-object
    # dependency) -- excluded here so they don't show up as a bogus parent
    # of themselves. Multiple lookups to the same parent object (or
    # multiple self-ref fields) are deduped with a set.
    parents_of = {}
    for e in edge_rows:
        if e["childobject"] == e["parentobject"]:
            continue
        if e["childobject"] in in_scope and e["parentobject"] in in_scope:
            parents_of.setdefault(e["childobject"], set()).add(e["parentobject"])

    level_members = {}
    for r in order_rows:
        level_members.setdefault(r["loadlevel"], []).append(r["objectname"])

    rows = []
    for r in order_rows:
        obj = r["objectname"]
        parents = sorted(parents_of.get(obj, []))
        siblings = sorted(n for n in level_members.get(r["loadlevel"], []) if n != obj)

        parts = [f"After: {', '.join(parents)}" if parents else "None"]
        if siblings:
            parts.append(f"parallel with: {', '.join(siblings)}")

        # Best-effort only -- no BulkOpsLog row exists yet at generation
        # time (no load has run), so there's no real SourceTable signal to
        # disambiguate a same-object-name, multiple-script case the way
        # _apply_log_result() below can once real load data exists. A
        # placeholder row built here that guesses the wrong one of two
        # real scripts gets corrected automatically the first time
        # sync_run_book_from_log() runs against a real log row for it.
        filename = sn.script_filename_for(obj, _TRANSFORMS_DIR, known_objects=in_scope)
        row = {
            "Stage": "Load",
            "Object": filename or obj,
            "Dependency": "; ".join(parts),
            "Status": "Not Started",
        }
        if filename and github_url:
            row["__hyperlink__"] = f'{github_url}/blob/{git_info["commit_sha"]}/sql/transformations/{filename}'
        rows.append(row)
    return rows


def _write_phase(ws, row, phase, load_rows):
    """Write one phase: a single banner row, then its data rows (no
    repeated column header, no blank spacer -- matching the real file's
    tight layout where the next banner immediately follows the last data
    row). A phase named "Load..." gets load_rows instead of its own
    template rows, when load_rows is available."""
    is_load_phase = phase["name"].lower().startswith("load")
    rows = load_rows if (is_load_phase and load_rows is not None) else phase["rows"]

    _write_banner(ws, row, phase["name"])
    row += 1
    for row_data in rows:
        _write_data_row(ws, row, row_data)
        row += 1
    return row


def generate_migration_run_book(output_path, tab_name, template_path=_TEMPLATE_PATH,
                                 engine=None, object_names=None, schema="dbo",
                                 project_name=None, source_env=None, target_env=None,
                                 ticket_url=None, ticket_label="JIRA"):
    """Create a brand-new Migration Run Book tab from
    docs/MIGRATION_RUN_BOOK_TEMPLATE.md (or a custom template_path).
    Refuses to overwrite an existing tab -- unlike mapping_doc.py's
    regenerate-in-place convention, a Migration Run Book tab holds live,
    manually-entered operational history that must never be silently
    clobbered.

    If object_names + engine are given, auto-fills the Load phase from
    analyze-load-order's (#2) results (with real hyperlinks to the matched
    script at the current commit, when this repo has a GitHub remote);
    omit them to fill that phase in by hand.

    project_name/source_env/target_env/ticket_url/ticket_label populate
    the fixed breadcrumb block (see module docstring) -- Git Repository/
    Commit-Branch/Scripts-link are always computed fresh from this repo's
    actual git state, not passed in."""
    phases = _parse_template(template_path)

    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        if tab_name in wb.sheetnames:
            raise ValueError(
                f"Tab '{tab_name}' already exists in {output_path} -- refusing to "
                "overwrite live Migration Run Book data. Pick a different --tab name."
            )
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    ws = wb.create_sheet(tab_name)

    git_info = gi.get_git_info()
    _write_breadcrumb_header(
        ws, git_info, project_name=project_name, source_env=source_env,
        target_env=target_env, ticket_url=ticket_url, ticket_label=ticket_label,
    )
    _write_table_header(ws, _TABLE_HEADER_ROW)

    load_rows = None
    if object_names:
        if engine is None:
            raise ValueError("object_names given without an engine")
        load_rows = _load_order_rows(engine, list(object_names), schema, git_info=git_info)

    row = _FIRST_DATA_ROW
    for phase in phases:
        row = _write_phase(ws, row, phase, load_rows)

    _apply_conditional_formatting(ws)
    _size_columns(ws)
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    _save_workbook(wb, output_path)
    return output_path


def _blank_result_columns_and_refresh(ws, git_info):
    """Blank every result column for each real data row (a banner row has
    no Object value, so it's skipped automatically), resetting Status to
    "Not Started" rather than leaving it empty -- a fresh pass's steps
    genuinely haven't started. Also re-derives each row's Object hyperlink
    against the *current* commit, working off whatever filename text is
    already there, including one a human typed in by hand."""
    recipe = set(_RECIPE_COLUMNS)
    result_idxs = [i for i, c in enumerate(_COLUMNS, start=1) if c not in recipe]
    status_idx = _COLUMNS.index("Status") + 1
    object_idx = _COLUMNS.index("Object") + 1
    github_url = gi.github_url(git_info["remote_url"]) if git_info else None

    for row_idx in range(_FIRST_DATA_ROW, ws.max_row + 1):
        object_value = ws.cell(row=row_idx, column=object_idx).value
        if not object_value:
            continue  # banner or blank row -- nothing to blank/refresh

        for col_idx in result_idxs:
            # openpyxl's cell(value=None) is a no-op (indistinguishable
            # from omitting value entirely) -- must set .value directly.
            ws.cell(row=row_idx, column=col_idx).value = None
        ws.cell(row=row_idx, column=status_idx).value = "Not Started"

        if github_url and os.path.isfile(os.path.join(_TRANSFORMS_DIR, str(object_value))):
            cell = ws.cell(row=row_idx, column=object_idx)
            cell.hyperlink = f'{github_url}/blob/{git_info["commit_sha"]}/sql/transformations/{object_value}'
            cell.font = _HYPERLINK_FONT


def add_migration_run_book_pass(path, from_tab, to_tab, template_path=_TEMPLATE_PATH,
                                 project_name=None, source_env=None, target_env=None,
                                 ticket_url=None, ticket_label=None):
    """Duplicate from_tab into a new to_tab: recipe columns (Stage/Object/
    Dependency/Critical/JIRA Ticket Link) carry forward verbatim, including
    any rows a human added by hand since generation; every result column
    is blanked for the fresh pass (Status reset to "Not Started"). Refuses
    to overwrite an existing to_tab.

    Breadcrumb carry-forward: Project/Source Environment/ticket URL-label
    carry forward from from_tab's own header unless explicitly overridden
    here (a project's Git repo and ticket-system project don't change pass
    to pass). Commit/Branch and the Scripts-link are always recomputed to
    the *current* git state. Target Environment is never silently carried
    forward -- Dev/UAT/PROD are different Salesforce orgs; pass it
    explicitly or it's left blank for a human to fill in."""
    wb = openpyxl.load_workbook(path)
    if from_tab not in wb.sheetnames:
        raise ValueError(f"No tab named '{from_tab}' in {path}")
    if to_tab in wb.sheetnames:
        raise ValueError(
            f"Tab '{to_tab}' already exists in {path} -- refusing to overwrite live Migration Run Book data."
        )

    src = wb[from_tab]
    carried = _read_breadcrumb_header(src)
    dst = wb.copy_worksheet(src)
    dst.title = to_tab

    git_info = gi.get_git_info()
    _write_breadcrumb_header(
        dst, git_info,
        project_name=project_name if project_name is not None else carried["project_name"],
        source_env=source_env if source_env is not None else carried["source_env"],
        target_env=target_env,
        ticket_url=ticket_url if ticket_url is not None else carried["ticket_url"],
        ticket_label=ticket_label if ticket_label is not None else carried["ticket_label"],
    )

    _blank_result_columns_and_refresh(dst, git_info)
    _apply_conditional_formatting(dst)

    _save_workbook(wb, path)
    return path


def _iter_load_phase_rows(ws, phase_prefix="load"):
    """Yield (row_idx, object_value) for every real data row currently
    belonging to a phase whose name starts with phase_prefix (case-
    insensitive) -- the same scoping _write_phase() uses to decide which
    phase gets load_order.py's auto-fill. Default "load" matches "Load
    Steps"; "pre" matches "Pre-Migration Steps" (roadmap #46's source-
    ingestion sync). A banner row (Object blank, Stage/label cell set)
    updates which phase subsequent rows belong to; it's never itself
    yielded."""
    object_col = _COLUMNS.index("Object") + 1
    current_phase = None
    row = _FIRST_DATA_ROW
    while row <= ws.max_row:
        object_value = ws.cell(row=row, column=object_col).value
        if not object_value:
            label = ws.cell(row=row, column=1).value
            if label:
                current_phase = label
            row += 1
            continue
        if current_phase and current_phase.lower().startswith(phase_prefix):
            yield row, object_value
        row += 1


_UNRESOLVED_STATUS_VALUES = {None, "", "Not Started"}


def _object_matches(object_name, object_value):
    """Does a row's Object cell refer to this Salesforce object? Delegates
    to script_numbering.matches_token() -- the same whole-token match this
    docstring used to duplicate verbatim (e.g. "Account" matches
    "010_account_load.sql" but "Order" does NOT match
    "030_orderitem_load.sql"). Kept as a single shared implementation so a
    future fix to the matching rule can't fix one call site and silently
    leave the other on the old, buggier logic."""
    return sn.matches_token(object_name, object_value)


def _find_pending_load_row(ws, object_name, phase_prefix="load"):
    """A phase_prefix-matching row whose Object refers to object_name (see
    _object_matches), whose Status is still unresolved (blank or the
    auto-fill default "Not Started" -- _load_order_rows() sets that
    explicitly, it isn't blank text), and whose Total Records is blank --
    an unresolved auto-fill placeholder, safe to fill in. A row with a
    real Status (Completed/Issue/In Process/N/A) or Total Records already
    set is either already resolved (a prior sync) or a human's own
    in-progress entry -- never touched. An exact-name match is preferred
    over a token match when both exist. Returns the row index, or None."""
    status_col = _COLUMNS.index("Status") + 1
    records_col = _COLUMNS.index("Total Records") + 1
    token_match = None
    for row_idx, object_value in _iter_load_phase_rows(ws, phase_prefix=phase_prefix):
        if not _object_matches(object_name, object_value):
            continue
        if ws.cell(row=row_idx, column=status_col).value not in _UNRESOLVED_STATUS_VALUES:
            continue
        if ws.cell(row=row_idx, column=records_col).value:
            continue
        if object_name.lower() == str(object_value).strip().lower():
            return row_idx
        if token_match is None:
            token_match = row_idx
    return token_match


def _insert_load_row(ws, object_name, phase_prefix="load"):
    """Insert a brand-new row for object_name right after the
    phase_prefix-matching phase's last existing row (before whatever phase
    follows it) -- never overwrites anything, used when no pending
    placeholder matches (e.g. a retry, or an object/file that was never
    pre-populated). Raises if this tab has no matching phase at all --
    nowhere safe to insert into."""
    object_col = _COLUMNS.index("Object") + 1
    last_match_row = None
    banner_row = None
    current_phase = None
    row = _FIRST_DATA_ROW
    while row <= ws.max_row:
        object_value = ws.cell(row=row, column=object_col).value
        if not object_value:
            label = ws.cell(row=row, column=1).value
            if label:
                current_phase = label
                if current_phase.lower().startswith(phase_prefix):
                    banner_row = row
            row += 1
            continue
        if current_phase and current_phase.lower().startswith(phase_prefix):
            last_match_row = row
        row += 1

    if last_match_row is not None:
        insert_at = last_match_row + 1
    elif banner_row is not None:
        insert_at = banner_row + 1
    else:
        raise ValueError(
            f"No phase starting with '{phase_prefix}' found in this tab -- add one "
            "(e.g. via docs/MIGRATION_RUN_BOOK_TEMPLATE.md) before syncing."
        )

    ws.insert_rows(insert_at)
    ws.cell(row=insert_at, column=object_col, value=object_name)
    return insert_at


def _apply_log_result(ws, row_idx, log_row, git_info=None, known_objects=None):
    # log_row's keys are already lowercased -- its one caller,
    # sync_run_book_from_log(), lowercases the whole batch once before
    # looping (found in review: this function used to re-lower here too,
    # a redundant second pass over the same row on every call). Postgres
    # returns an unquoted column lowercased in its own query results,
    # unlike SQL Server/SQLite -- see sql_dialect.lower_keys()'s docstring.
    submitted = log_row["recordssubmitted"]
    succeeded = log_row["recordssucceeded"]
    failed = log_row["recordsfailed"]

    values = {
        "Status": "Issue" if failed else "Completed",
        "Person Responsible": log_row["runby"],
        "Begin Time": log_row["startedat"],
        "End Time": log_row["completedat"],
        "Total Records": submitted,
        "Success Records": succeeded,
        "Failed Records": failed,
        "Notes": f"Auto-synced from {log_row['targetschema']}.BulkOpsLog #{log_row['logid']} "
                 f"({log_row['operation']}, {log_row['jobcount']} job(s)).",
    }
    for col_name, value in values.items():
        col_idx = _COLUMNS.index(col_name) + 1
        ws.cell(row=row_idx, column=col_idx, value=value)

    if submitted:
        pct_idx = _COLUMNS.index("Success Percent") + 1
        cell = ws.cell(row=row_idx, column=pct_idx, value=succeeded / submitted)
        cell.number_format = "0.00%"

    # Prefer the real transform script's filename over a plain object name
    # in the Object column -- the placeholder row created at generation
    # time may predate the script actually being built (or may have
    # matched a stale/illustrative script that has since been replaced),
    # so this is re-resolved on every sync rather than trusted from
    # whatever's already in the cell.
    #
    # When more than one real script implements this same object name from
    # different source-routing branches (e.g. this project's own
    # GiftTransaction: 200 from Opportunity, 210 from Payment),
    # script_filename_for()'s own highest-number-wins tiebreak can silently
    # pick the wrong one -- found live, 2026-07-18, syncing this exact
    # object after a corrective reload against the Opportunity branch
    # linked to 210 (the Payment branch) instead of 200. SourceTable is a
    # real, unambiguous signal here (each script's own SELECT INTO target
    # is exactly what bulk_op() recorded there), so it's tried first via
    # script_filename_for_source_table(); the original highest-number
    # default only applies when that can't disambiguate (a single
    # candidate, or SourceTable not matching any of them).
    #
    # Known residual gap, not fixed here: a `delete` operation's own
    # SourceTable is always a `<Object>_Purge` table (bulkops.py's
    # purge_by_filter()), never one of the real scripts' own build tables
    # -- there's no per-branch signal to disambiguate a delete row this
    # way at all, so a multi-script object's delete rows still fall back
    # to the highest-numbered guess, same as before this fix.
    candidates = sn.script_candidates_for(log_row["objectname"], _TRANSFORMS_DIR, known_objects=known_objects)
    filename = ""
    if len(candidates) > 1:
        filename = sn.script_filename_for_source_table(
            candidates, _TRANSFORMS_DIR, log_row.get("sourcetable")
        )
    if not filename:
        filename = candidates[-1] if candidates else ""
    if filename:
        object_idx = _COLUMNS.index("Object") + 1
        cell = ws.cell(row=row_idx, column=object_idx, value=filename)
        github_url = gi.github_url(git_info["remote_url"]) if git_info else None
        if github_url:
            cell.hyperlink = f'{github_url}/blob/{git_info["commit_sha"]}/sql/transformations/{filename}'
            cell.font = _HYPERLINK_FONT


def sync_run_book_from_log(engine, path, tab_name, schema="dbo"):
    """Pull dbo.BulkOpsLog (#14) rows not yet reflected in this tab into
    its Load phase: fills in a still-pending auto-fill placeholder row for
    that object if one exists, otherwise inserts a new row (e.g. a retry)
    -- never overwrites a row that already has real result data, and never
    touches anything outside the Load phase. Tracks a per-tab watermark
    (Last Synced Log Id, in the breadcrumb block) so re-running only ever
    pulls in genuinely new log entries -- safe to call repeatedly, and
    safe to call after a human has added their own rows by hand.

    v1 limit: only pulls dbo.BulkOpsLog's own aggregate columns (record
    counts, timing) -- per-row Error Details text would need reading the
    separate `_Result` writeback table too, not done here."""
    wb = openpyxl.load_workbook(path)
    if tab_name not in wb.sheetnames:
        raise ValueError(f"No tab named '{tab_name}' in {path}")
    ws = wb[tab_name]

    last_synced = ws.cell(row=_HEADER_ROW_LAST_SYNCED_LOG_ID, column=2).value or 0

    d = sql_dialect.for_engine(engine)
    if not d.table_exists(engine, schema, "BulkOpsLog"):
        return {"synced": 0, "inserted": 0, "updated": 0,
                "message": f"{schema}.BulkOpsLog doesn't exist yet -- nothing to sync."}

    with engine.connect() as cx:
        new_rows = cx.execute(
            text(
                f"SELECT LogId, ObjectName, SourceTable, Operation, TargetSchema, RecordsSubmitted, "
                f"RecordsSucceeded, RecordsFailed, StartedAt, CompletedAt, RunBy, JobCount "
                f"FROM {d.qualify(schema, 'BulkOpsLog')} WHERE LogId > :last ORDER BY LogId"
            ),
            {"last": last_synced},
        ).mappings().all()

    if not new_rows:
        return {"synced": 0, "inserted": 0, "updated": 0}

    # Normalized to lowercase keys once here -- see _apply_log_result()'s
    # own docstring for why (Postgres returns an unquoted column
    # lowercased in its own query results, unlike SQL Server/SQLite).
    new_rows = [sql_dialect.lower_keys(r) for r in new_rows]

    # The distinct object names in this sync batch -- passed to
    # script_filename_for() as known_objects so a compound-name script
    # (e.g. AccountContactRelation) can't silently outrank the real
    # script for a shorter object it happens to embed (e.g. Account),
    # the same collision found live in the NPSP-to-NPC migration
    # (ROADMAP #76). Scoped to this batch, not a full project registry --
    # still a real, partial improvement over no known_objects at all.
    known_objects = {r["objectname"] for r in new_rows}

    git_info = gi.get_git_info()
    inserted, updated = 0, 0
    for log_row in new_rows:
        target_row = _find_pending_load_row(ws, log_row["objectname"])
        if target_row is None:
            target_row = _insert_load_row(ws, log_row["objectname"])
            inserted += 1
        else:
            updated += 1
        _apply_log_result(ws, target_row, log_row, git_info=git_info, known_objects=known_objects)

    ws.cell(row=_HEADER_ROW_LAST_SYNCED_LOG_ID, column=2, value=new_rows[-1]["logid"])
    _save_workbook(wb, path)
    return {"synced": len(new_rows), "inserted": inserted, "updated": updated}


def _apply_source_ingestion_result(ws, row_idx, log_row, schema):
    # log_row's keys are already lowercased by its one caller,
    # sync_source_ingestion_to_run_book() -- see _apply_log_result()'s
    # own comment for why this used to (redundantly) re-lower here too.
    row_count = log_row["rowcount"]
    is_blocked = log_row["status"] == "drift_blocked"

    values = {
        "Status": "Issue" if is_blocked else "Completed",
        "Person Responsible": log_row["runby"],
        "Begin Time": log_row["startedat"],
        "End Time": log_row["completedat"],
        "Notes": f"Auto-synced from {schema}.SourceIngestionLog #{log_row['logid']} ({log_row['status']}).",
    }
    if not is_blocked:
        values["Total Records"] = row_count
        values["Success Records"] = row_count
        values["Failed Records"] = 0
    if is_blocked and log_row["driftdetails"]:
        values["Error Details"] = log_row["driftdetails"]

    for col_name, value in values.items():
        col_idx = _COLUMNS.index(col_name) + 1
        ws.cell(row=row_idx, column=col_idx, value=value)

    if not is_blocked and row_count:
        pct_idx = _COLUMNS.index("Success Percent") + 1
        cell = ws.cell(row=row_idx, column=pct_idx, value=1.0)
        cell.number_format = "0.00%"


def sync_source_ingestion_to_run_book(engine, path, tab_name, schema="dbo"):
    """Pull dbo.SourceIngestionLog (roadmap #46) rows not yet reflected in
    this tab into its Pre-Migration phase (matched via phase_prefix="pre")
    -- same fill-placeholder-or-insert, watermarked, idempotent mechanism
    as sync_run_book_from_log(), just targeting a different log table and
    phase, with its own independent watermark
    (_HEADER_ROW_LAST_SYNCED_SOURCE_LOG_ID) so the two syncs never
    interfere with each other. A drift-blocked run is still logged (see
    source_ingestion.import_directory) and lands here as Status="Issue"
    with the exact column diff in Error Details -- visible in the audit
    trail, not just a console message."""
    wb = openpyxl.load_workbook(path)
    if tab_name not in wb.sheetnames:
        raise ValueError(f"No tab named '{tab_name}' in {path}")
    ws = wb[tab_name]

    last_synced = ws.cell(row=_HEADER_ROW_LAST_SYNCED_SOURCE_LOG_ID, column=2).value or 0

    d = sql_dialect.for_engine(engine)
    if not d.table_exists(engine, schema, "SourceIngestionLog"):
        return {"synced": 0, "inserted": 0, "updated": 0,
                "message": f"{schema}.SourceIngestionLog doesn't exist yet -- nothing to sync."}

    with engine.connect() as cx:
        new_rows = cx.execute(
            text(
                f"SELECT LogId, TableName, Status, {d.quote_ident('RowCount')}, StartedAt, CompletedAt, RunBy, DriftDetails "
                f"FROM {d.qualify(schema, 'SourceIngestionLog')} WHERE LogId > :last ORDER BY LogId"
            ),
            {"last": last_synced},
        ).mappings().all()

    if not new_rows:
        return {"synced": 0, "inserted": 0, "updated": 0}

    # Normalized to lowercase keys once here -- see _apply_log_result()'s
    # own docstring for why (Postgres returns an unquoted column
    # lowercased in its own query results, unlike SQL Server/SQLite).
    new_rows = [sql_dialect.lower_keys(r) for r in new_rows]

    inserted, updated = 0, 0
    for log_row in new_rows:
        target_row = _find_pending_load_row(ws, log_row["tablename"], phase_prefix="pre")
        if target_row is None:
            target_row = _insert_load_row(ws, log_row["tablename"], phase_prefix="pre")
            inserted += 1
        else:
            updated += 1
        _apply_source_ingestion_result(ws, target_row, log_row, schema)

    ws.cell(row=_HEADER_ROW_LAST_SYNCED_SOURCE_LOG_ID, column=2, value=new_rows[-1]["logid"])
    _save_workbook(wb, path)
    return {"synced": len(new_rows), "inserted": inserted, "updated": updated}


# --- Mermaid process-flow diagram (roadmap #52) --------------------------
#
# Deliberately simple for v1, per the user's own framing ("for now, you
# would just create mermaid"): emit the Mermaid syntax itself off this
# tab's own Stage/Object/Dependency/Status columns and stop there --
# GitHub/most Markdown renderers already render a fenced ```mermaid```
# block natively, and Lucid supports paste-to-import, so no further
# tooling is needed to get value from this. A *polished*, hand-styled
# diagram elsewhere is a named future stretch, not part of this item.

_DEPENDENCY_AFTER_RE = re.compile(r"^After:\s*(.+?)(?:\s*;\s*parallel with:.*)?$", re.IGNORECASE)

_STATUS_CSS_CLASS = {
    "Not Started": "statusNotStarted",
    "N/A": "statusNA",
    "In Process": "statusInProcess",
    "Completed": "statusCompleted",
    "Issue": "statusIssue",
}
# Same five-state palette as _STATUS_FILLS/_STATUS_FONTS above -- the
# flowchart's node coloring should visually agree with the workbook's own
# conditional formatting, not invent a separate meaning for color.
_STATUS_CLASS_DEFS = {
    "statusNotStarted": "fill:#FFFFFF,stroke:#333333,color:#000000",
    "statusNA": "fill:#C6EFCE,stroke:#375623,color:#000000",
    "statusInProcess": "fill:#FFFF00,stroke:#333333,color:#000000",
    "statusCompleted": "fill:#375623,stroke:#1f2937,color:#FFFFFF",
    "statusIssue": "fill:#FF0000,stroke:#333333,color:#FFFFFF",
}


def _parse_dependency_parents(dependency_text):
    """Extract the comma-separated parent names from a Dependency cell's
    "After: X, Y[; parallel with: ...]" text -- the exact shape
    _load_order_rows() itself writes. "None" or blank yields no parents.
    The "parallel with" suffix names siblings at the same load level, not
    a dependency, so it's never turned into an edge."""
    if not dependency_text or str(dependency_text).strip().lower() == "none":
        return []
    m = _DEPENDENCY_AFTER_RE.match(str(dependency_text).strip())
    if not m:
        return []
    return [p.strip() for p in m.group(1).split(",") if p.strip()]


def _is_unparseable_dependency_note(dependency_text):
    """True when dependency_text has real content but doesn't match the
    "After: X" shape _parse_dependency_parents() understands -- e.g. a
    human typed a free-text note directly into the Dependency cell
    instead of using that convention (plausible: a Pre-/Post-Migration
    row's Dependency column is hand-filled, not auto-generated). Distinct
    from blank/"None" (which genuinely means no dependency) -- found in
    review: both cases previously looked identical to
    generate_run_book_flowchart() (an empty parent list from
    _parse_dependency_parents() either way), so a real dependency note a
    human wrote in free text was silently dropped from the diagram with
    no signal at all that anything was skipped."""
    if not dependency_text or str(dependency_text).strip().lower() == "none":
        return False
    return _DEPENDENCY_AFTER_RE.match(str(dependency_text).strip()) is None


def _mermaid_escape_label(text):
    """Escape a label for Mermaid's ["..."] node/subgraph syntax -- an
    embedded double-quote or bracket would otherwise break out of the
    node shape and corrupt the diagram."""
    return str(text).replace('"', "#quot;").replace("[", "(").replace("]", ")")


def generate_run_book_flowchart(path, tab_name):
    """Generate a Mermaid flowchart straight from a Migration Run Book
    tab's own structure: one subgraph per phase banner, one node per data
    row (labeled from the Object column), edges parsed from the
    Dependency column's "After: X, Y" text, and node color matching the
    workbook's own Status conditional-formatting palette.

    Parent names are resolved against every OTHER row's Object value in
    this tab via script_numbering.matches_token() -- the same whole-token
    match already used to resolve a real script filename against a bare
    object name elsewhere in this module, since a Load-phase Object cell
    is often a script filename ("010_account_load.sql"), not the bare
    name a Dependency cell names ("After: Account"). A dependency name
    with no matching row in this tab is dropped rather than guessed at --
    returned in the summary dict's "unresolved_dependencies" list so a
    caller can surface it, same "visible gap, not a silent guess"
    philosophy as record_types.py's unmatched-DeveloperName NULL.

    Returns (mermaid_text, summary_dict) -- mermaid_text is a complete
    fenced ```mermaid code block, ready to write straight to a .md file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if tab_name not in wb.sheetnames:
        raise ValueError(f"No tab named '{tab_name}' in {path}")
    ws = wb[tab_name]

    object_col = _COLUMNS.index("Object") + 1
    dependency_col = _COLUMNS.index("Dependency") + 1
    status_col = _COLUMNS.index("Status") + 1

    phase_order = []
    nodes_by_phase = {}
    current_phase = None
    node_counter = 0

    row = _FIRST_DATA_ROW
    while row <= ws.max_row:
        object_value = ws.cell(row=row, column=object_col).value
        if not object_value:
            label = ws.cell(row=row, column=1).value
            if label:
                current_phase = label
                if current_phase not in nodes_by_phase:
                    nodes_by_phase[current_phase] = []
                    phase_order.append(current_phase)
            row += 1
            continue
        node_counter += 1
        status = ws.cell(row=row, column=status_col).value or "Not Started"
        dependency_text = ws.cell(row=row, column=dependency_col).value
        nodes_by_phase.setdefault(current_phase, [])
        if current_phase not in phase_order:
            phase_order.append(current_phase)
        nodes_by_phase[current_phase].append({
            "id": f"n{node_counter}", "label": str(object_value),
            "status": status, "dependency_text": dependency_text,
        })
        row += 1

    all_nodes = [n for phase_nodes in nodes_by_phase.values() for n in phase_nodes]

    lines = ["```mermaid", "flowchart TD"]
    edges = []
    unresolved = []
    unparsed_notes = []
    for phase_name in phase_order:
        phase_nodes = nodes_by_phase.get(phase_name) or []
        if not phase_nodes:
            continue
        subgraph_id = re.sub(r"\W+", "_", phase_name).strip("_") or "phase"
        lines.append(f'    subgraph {subgraph_id}["{_mermaid_escape_label(phase_name)}"]')
        for n in phase_nodes:
            css_class = _STATUS_CSS_CLASS.get(n["status"], "statusNotStarted")
            lines.append(f'        {n["id"]}["{_mermaid_escape_label(n["label"])}"]:::{css_class}')
        lines.append("    end")

        for n in phase_nodes:
            if _is_unparseable_dependency_note(n["dependency_text"]):
                unparsed_notes.append(f'{n["label"]}: "{n["dependency_text"]}"')
            for parent_name in _parse_dependency_parents(n["dependency_text"]):
                match = next(
                    (other for other in all_nodes
                     if other is not n and sn.matches_token(parent_name, other["label"])),
                    None,
                )
                if match:
                    edges.append((match["id"], n["id"]))
                else:
                    unresolved.append(parent_name)

    for parent_id, child_id in edges:
        lines.append(f"    {parent_id} --> {child_id}")
    for class_name, style in _STATUS_CLASS_DEFS.items():
        lines.append(f"    classDef {class_name} {style}")
    lines.append("```")

    mermaid_text = "\n".join(lines)
    summary = {
        "phases": len([p for p in phase_order if nodes_by_phase.get(p)]),
        "nodes": len(all_nodes),
        "edges": len(edges),
        "unresolved_dependencies": unresolved,
        "unparsed_dependency_notes": unparsed_notes,
    }
    return mermaid_text, summary
