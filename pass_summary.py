"""Auto-drafted client-facing pass summary (roadmap #66).

A plain-English "here's what happened this pass" draft, pulled straight
from a Migration Run Book tab's own Load-phase rows (migration_run_book.py,
roadmap #16) -- object count, total/succeeded/failed records per object --
ready to send a client stakeholder instead of a raw spreadsheet dump or a
manually-written status email. Nearly free given how much structured data
this framework already logs by the time a pass finishes.

Plain Markdown for v1, not a Word document -- same "ship the simple
version, decide on polish later" discipline as #52's own v1 framing
(generate_run_book_flowchart()); solution_doc.py's docxtpl machinery is
there to reuse later if a client-ready Word format is ever wanted instead.

Optional, explicit failure-detail enrichment: pass load_tables (an
{object_name: sql_table_name} dict) to have any Load-phase row with
failed records > 0 cross-referenced through failure_triage.py (#61) for a
plain-language root cause instead of just a raw failed count. Never
guesses a SQL table name from a Run Book row's Object cell (which may be
a bare object name or a script filename, see migration_run_book.py's own
Object-column docstring) -- an object left out of load_tables just gets
a pointer at the Run Book's own Notes/Error Details columns instead.
"""
import openpyxl

import failure_triage as ft
from migration_run_book import _COLUMNS, _HEADER_ROW_PROJECT, _HEADER_ROW_TARGET_ENV, _iter_load_phase_rows


def _load_phase_summary_rows(run_book_path, tab_name):
    wb = openpyxl.load_workbook(run_book_path, data_only=True)
    if tab_name not in wb.sheetnames:
        raise ValueError(f"No tab named '{tab_name}' in {run_book_path}")
    ws = wb[tab_name]

    project = ws.cell(row=_HEADER_ROW_PROJECT, column=2).value
    target_env = ws.cell(row=_HEADER_ROW_TARGET_ENV, column=2).value

    status_col = _COLUMNS.index("Status") + 1
    total_col = _COLUMNS.index("Total Records") + 1
    success_col = _COLUMNS.index("Success Records") + 1
    failed_col = _COLUMNS.index("Failed Records") + 1

    rows = []
    for row_idx, object_value in _iter_load_phase_rows(ws, phase_prefix="load"):
        rows.append({
            "object": object_value,
            "status": ws.cell(row=row_idx, column=status_col).value,
            "total": ws.cell(row=row_idx, column=total_col).value,
            "succeeded": ws.cell(row=row_idx, column=success_col).value,
            "failed": ws.cell(row=row_idx, column=failed_col).value,
        })
    return project, target_env, rows


def generate_pass_summary(run_book_path, tab_name, engine=None, schema="dbo", load_tables=None):
    """Draft a plain-English pass summary from run_book_path's tab_name
    Load phase. Returns the summary as a Markdown string.

    load_tables: optional {object_name: sql_table_name} -- for any
    Load-phase row with failed records > 0 whose object appears here (and
    engine is given), calls failure_triage.triage_failures() against that
    table for a plain-language root cause per distinct error signature.
    An object with failures but no entry here (the common case unless
    explicitly wired up) falls back to pointing at the Run Book's own
    Notes/Error Details columns instead -- this never guesses which SQL
    table a Run Book row's Object cell (a bare object name, or a script
    filename like "020_contact_load.sql") actually corresponds to.
    """
    load_tables = load_tables or {}
    project, target_env, rows = _load_phase_summary_rows(run_book_path, tab_name)

    total_objects = len(rows)
    total_records = sum(r["total"] or 0 for r in rows)
    total_succeeded = sum(r["succeeded"] or 0 for r in rows)
    total_failed = sum(r["failed"] or 0 for r in rows)

    lines = [f"# Migration Pass Summary{f' -- {project}' if project else ''}"]
    if target_env:
        lines.append(f"\n**Target environment:** {target_env}")
    lines.append(
        f"\nThis pass migrated **{total_objects} object(s)** across "
        f"**{total_records} record(s)**: {total_succeeded} succeeded"
        + (f", {total_failed} had exception(s)" if total_failed else "")
        + "."
    )

    lines.append("\n## Per-object results\n")
    lines.append("| Object | Status | Total | Succeeded | Failed |")
    lines.append("|---|---|---|---|---|")
    for r in rows:
        lines.append(
            f"| {r['object'] or ''} | {r['status'] or ''} | {r['total'] or 0} | "
            f"{r['succeeded'] or 0} | {r['failed'] or 0} |"
        )

    issues = [r for r in rows if (r["failed"] or 0) > 0]
    if issues:
        lines.append("\n## Known issues\n")
        for r in issues:
            lines.append(f"### {r['object']}\n")
            lines.append(f"{r['failed']} of {r['total'] or 0} record(s) had an exception.")
            table_name = load_tables.get(r["object"])
            if table_name and engine is not None:
                try:
                    triage = ft.triage_failures(engine, table_name, schema=schema)
                except ValueError:
                    # triage_failures() raises when table_name has no Error
                    # column at all -- including when the table doesn't
                    # exist anymore (found in review: this is exactly the
                    # "cleaned up since this pass ran" scenario the
                    # fallback message below was meant to cover, but an
                    # uncaught raise never reached it).
                    triage = None
                if triage:
                    for t in triage:
                        lines.append(f"- **{t['code']}** ({t['count']}x): {t['cause']}")
                else:
                    lines.append(
                        "(No failure detail found in that load table -- it may have "
                        "been cleaned up or reloaded since this pass ran.)"
                    )
            else:
                lines.append("(See the Migration Run Book's Notes/Error Details columns for more detail.)")
    else:
        lines.append("\nNo exceptions this pass.")

    return "\n".join(lines) + "\n"
