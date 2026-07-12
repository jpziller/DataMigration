"""Row-count reconciliation report (roadmap #64).

A data-completeness auditor spanning the whole load order, not a
per-tool spot check: cross-checks three numbers that should reconcile
for each object -- the source table's row count, the Load table's row
count (did the transform's JOINs/WHERE clauses silently drop rows it
shouldn't have?), and bulkops' most recent submitted/succeeded/failed
counts from BulkOpsLog (#14). Entirely read-only, aggregating data every
one of these tools already produces -- the value is in cross-checking
all three together in one pass, not new data collection.

Source table discovery: reads a mapping doc's own "Source Object:"
header cell (row 1, the exact cell generate_mapping_workbook() writes)
for each object's sheet, when mapping_path is given -- never guessed.
Without a mapping doc, the source-table half of the reconciliation is
simply skipped for that object (still reports Load/bulkops numbers).

Load table naming: defaults to "<object_name>_Load", matching this
project's own overwhelming convention (same default reset-dev-cycle/
pass_summary.py already use) -- override per object via load_tables for
a project that deviates from it.
"""
import openpyxl
from sqlalchemy import text

import sql_dialect
from mapping_doc import _safe_sheet_name


def _source_table_from_mapping(mapping_path, object_name):
    """The source table name recorded in object_name's mapping-doc sheet
    header (row 1, col 2 -- the "Source Object:" cell
    generate_mapping_workbook() writes) -- None if there's no mapping doc,
    no sheet for this object, or the header cell is blank."""
    if not mapping_path:
        return None
    try:
        wb = openpyxl.load_workbook(mapping_path, data_only=True)
    except FileNotFoundError:
        return None
    sheet_name = _safe_sheet_name(object_name)
    if sheet_name not in wb.sheetnames:
        return None
    value = wb[sheet_name].cell(row=1, column=2).value
    return str(value).strip() if value else None


def _row_count(engine, d, schema, table_name):
    if not d.table_exists(engine, schema, table_name):
        return None
    with engine.connect() as cx:
        return cx.execute(text(f"SELECT COUNT(*) FROM {d.qualify(schema, table_name)}")).scalar()


def _latest_bulkops_row(engine, d, object_name, schema="dbo"):
    if not d.table_exists(engine, schema, "BulkOpsLog"):
        return None
    query = d.select_top_n_sql(
        "RecordsSubmitted, RecordsSucceeded, RecordsFailed",
        f"FROM {d.qualify(schema, 'BulkOpsLog')} WHERE ObjectName = :obj ORDER BY LogId DESC",
        1,
    )
    with engine.connect() as cx:
        row = cx.execute(text(query), {"obj": object_name}).mappings().first()
    return dict(row) if row else None


def reconcile_load_counts(engine, object_names, schema="dbo", mapping_path=None, load_tables=None):
    """Reconcile source/Load/bulkops row counts for each name in
    object_names. Returns [{"object", "source_table", "source_count",
    "load_table", "load_count", "bulkops_submitted", "bulkops_succeeded",
    "bulkops_failed", "flags"}, ...] in the order given -- "flags" is a
    list of plain-English strings for anything that doesn't reconcile the
    way it's supposed to; empty means clean."""
    d = sql_dialect.for_engine(engine)
    load_tables = load_tables or {}
    results = []

    for object_name in object_names:
        load_table = load_tables.get(object_name, f"{object_name}_Load")
        load_count = _row_count(engine, d, schema, load_table)

        source_table = _source_table_from_mapping(mapping_path, object_name)
        source_count = _row_count(engine, d, schema, source_table) if source_table else None

        bulkops_row = _latest_bulkops_row(engine, d, object_name, schema=schema)

        flags = []
        if load_count is None:
            flags.append(f"Load table {schema}.{load_table} doesn't exist yet.")
        if source_count is not None and load_count is not None and load_count < source_count:
            flags.append(
                f"Load table has fewer rows than source ({load_count} vs {source_count}) -- "
                f"the transform may have dropped {source_count - load_count} row(s)."
            )
        if bulkops_row is None and load_count is not None:
            flags.append("Never loaded via bulkops yet (no BulkOpsLog row for this object).")
        if bulkops_row is not None and load_count is not None and bulkops_row["RecordsSubmitted"] != load_count:
            flags.append(
                f"Load table now has {load_count} row(s), but the most recent bulkops run submitted "
                f"{bulkops_row['RecordsSubmitted']} -- this may reflect a stale prior run; rerun bulkops "
                "to pick up the current Load table."
            )

        results.append({
            "object": object_name,
            "source_table": source_table, "source_count": source_count,
            "load_table": load_table, "load_count": load_count,
            "bulkops_submitted": bulkops_row["RecordsSubmitted"] if bulkops_row else None,
            "bulkops_succeeded": bulkops_row["RecordsSucceeded"] if bulkops_row else None,
            "bulkops_failed": bulkops_row["RecordsFailed"] if bulkops_row else None,
            "flags": flags,
        })

    return results
