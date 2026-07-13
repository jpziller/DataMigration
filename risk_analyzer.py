"""Org metadata risk analyzer (roadmap #5).

Cross-references the objects a migration touches against the target org's
live automation metadata -- validation rules, Apex triggers, record-
triggered Flows, legacy Workflow Rules, and approval processes -- so a data
architect knows what could interfere with a load *before* running bulkops
for real, not after an unexplained rejection or a cascading side effect.

Scope, stated plainly: this is an OBJECT-level automation inventory, not a
field-level formula parser. Reliably determining exactly which fields a
validation rule's formula or a Flow's condition logic references would need
either brittle text-scanning of formula strings, or a much heavier Metadata
API retrieval per Flow -- deferred, the same way auto_mapper.py deferred its
data-sampling "layer 4" until real usage shows it's actually needed. One
concrete, honest field-level signal IS included cheaply: an active
validation rule's ErrorDisplayField (the one field Salesforce highlights
when the rule fires) is cross-referenced against whichever target fields
are actually being migrated for that object, if supplied -- see
`fields_in_scope` below.

Two different Salesforce API surfaces are involved here, and using the
wrong one for a given metadata type silently returns an error or an empty
result rather than the data you wanted -- worth being explicit about, and
verified live against a real org before shipping (this org runs a
post-training-cutoff API version -- see CLAUDE.md on not trusting
possibly-stale assumptions here):
  - Tooling API (`sf.toolingexecute`): ValidationRule, ApexTrigger, WorkflowRule
  - Standard REST Query API (`sf.query`): ProcessDefinition, FlowDefinitionView
"""
import os
import re
import urllib.parse
from datetime import datetime, timezone

import openpyxl
from sqlalchemy import text

import sql_dialect

_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")

# Mirrors mapping_doc.py's column layout -- see that module's _HEADERS.
_COL_MIGRATE_DATA = 9
_COL_TARGET_FIELD_API = 15


def _safe_sheet_name(name):
    return _INVALID_SHEET_CHARS.sub("_", name)[:31]


def fields_in_scope_from_mapping(mapping_path, object_names):
    """Read each object's mapping-doc sheet (if present) and return
    {object_name: {target field API names where Migrate Data == 'Yes'}}.
    An object with no mapping doc given, or no sheet for it, is simply
    absent from the returned dict -- analyze_object_risk treats that the
    same as "no scope info supplied," not an error."""
    if not mapping_path or not os.path.exists(mapping_path):
        return {}

    wb = openpyxl.load_workbook(mapping_path, data_only=True)
    result = {}
    for name in object_names:
        sheet_name = _safe_sheet_name(name)
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        fields = set()
        for row in ws.iter_rows(min_row=4):
            migrate = row[_COL_MIGRATE_DATA - 1].value
            target_field = row[_COL_TARGET_FIELD_API - 1].value
            if migrate and str(migrate).strip() == "Yes" and target_field:
                fields.add(str(target_field).strip())
        result[name] = fields
    return result


def _tooling_query(sf, soql):
    return sf.toolingexecute(f"query/?q={urllib.parse.quote(soql)}")


def _validation_rules(sf, object_name):
    r = _tooling_query(sf, (
        "SELECT Id, ValidationName, Active, ErrorDisplayField, ErrorMessage, Description "
        "FROM ValidationRule "
        f"WHERE EntityDefinition.QualifiedApiName = '{object_name}'"
    ))
    return r.get("records", [])


def _apex_triggers(sf, object_name):
    r = _tooling_query(sf, (
        "SELECT Id, Name, UsageBeforeInsert, UsageAfterInsert, UsageBeforeUpdate, "
        "UsageAfterUpdate, UsageBeforeDelete, UsageAfterDelete, UsageAfterUndelete "
        "FROM ApexTrigger "
        f"WHERE EntityDefinition.QualifiedApiName = '{object_name}'"
    ))
    return r.get("records", [])


def _workflow_rules(sf, object_name):
    r = _tooling_query(sf, f"SELECT Id, Name FROM WorkflowRule WHERE TableEnumOrId = '{object_name}'")
    return r.get("records", [])


def _approval_processes(sf, object_name):
    r = sf.query(f"SELECT Id, Name, State FROM ProcessDefinition WHERE TableEnumOrId = '{object_name}'")
    return r.get("records", [])


def _record_triggered_flows(sf, object_name):
    r = sf.query(
        "SELECT Id, Label, ProcessType, TriggerType, IsActive FROM FlowDefinitionView "
        f"WHERE TriggerObjectOrEventLabel = '{object_name}' "
        "AND TriggerType IN ('RecordBeforeSave', 'RecordAfterSave')"
    )
    return r.get("records", [])


def analyze_object_risk(sf, object_name, fields_in_scope=None):
    """Return an automation-risk inventory for one object.

    fields_in_scope, if given, is an iterable of target field API names
    actually being migrated for this object (e.g. a mapping doc's Target
    Field API column where Migrate Data == 'Yes'). It's used only to flag
    an active validation rule's ErrorDisplayField as a direct hit -- it
    doesn't filter or limit anything else in the report.
    """
    warnings = []
    scope = set(fields_in_scope or [])

    def _safe(fn, label):
        try:
            return fn(sf, object_name)
        except Exception as e:
            warnings.append(f"Could not check {label} for {object_name}: {e}")
            return []

    validation_rules = _safe(_validation_rules, "validation rules")
    apex_triggers = _safe(_apex_triggers, "Apex triggers")
    workflow_rules = _safe(_workflow_rules, "workflow rules")
    approval_processes = _safe(_approval_processes, "approval processes")
    record_triggered_flows = _safe(_record_triggered_flows, "record-triggered flows")

    active_rules = [r for r in validation_rules if r.get("Active")]
    for r in active_rules:
        r["direct_hit"] = bool(scope) and r.get("ErrorDisplayField") in scope

    return {
        "object": object_name,
        "validation_rules": validation_rules,
        "active_validation_rule_count": len(active_rules),
        "direct_hit_validation_rules": [r for r in active_rules if r["direct_hit"]],
        "apex_triggers": apex_triggers,
        "workflow_rules": workflow_rules,
        "approval_processes": approval_processes,
        "record_triggered_flows": record_triggered_flows,
        "active_flow_count": sum(1 for f in record_triggered_flows if f.get("IsActive")),
        "warnings": warnings,
    }


def analyze_migration_risk(sf, object_names, fields_in_scope_by_object=None):
    fields_in_scope_by_object = fields_in_scope_by_object or {}
    return [
        analyze_object_risk(sf, name, fields_in_scope=fields_in_scope_by_object.get(name))
        for name in object_names
    ]


def _ensure_table(engine, schema="dbo"):
    d = sql_dialect.for_engine(engine)
    if d.table_exists(engine, schema, "ObjectAutomationRisk"):
        return
    qualified = d.qualify(schema, "ObjectAutomationRisk")
    with engine.begin() as cx:
        # Column names here are deliberately bare (not d.quote_ident()) --
        # found via live Postgres testing: quoting at CREATE TABLE
        # preserves exact case in Postgres's catalog, but every read/
        # write of this table elsewhere (write_to_sql()'s own DELETE/
        # INSERT below, plus batch_advisor.py's/failure_triage.py's/
        # orchestrator.py's/readiness.py's own reads, none of which quote
        # their references) uses bare column references, which Postgres
        # folds to lowercase -- bare here matches that dominant
        # convention instead of requiring every scattered reference to be
        # quoted.
        cx.execute(text(
            f"CREATE TABLE {qualified} ("
            f"ObjectName {d.pick_type('NVARCHAR(255)', 'TEXT', 'VARCHAR(255)')} NOT NULL, "
            f"CheckType {d.pick_type('NVARCHAR(50)', 'TEXT', 'VARCHAR(50)')} NOT NULL, "
            f"ItemName {d.pick_type('NVARCHAR(255)', 'TEXT', 'VARCHAR(255)')} NOT NULL, "
            f"IsActive {d.pick_type('BIT', 'INTEGER', 'BOOLEAN')} NULL, "
            f"DirectHit {d.pick_type('BIT', 'INTEGER', 'BOOLEAN')} NULL, "
            f"Detail {d.raw_text_type()} NULL, "
            f"AnalyzedDate {d.pick_type('DATETIME2', 'TEXT', 'TIMESTAMP')} NOT NULL);"
        ))


def write_to_sql(engine, results, schema="dbo"):
    _ensure_table(engine, schema=schema)
    analyzed_at = datetime.now(timezone.utc).replace(tzinfo=None)

    rows = []
    for result in results:
        obj = result["object"]
        rows_before = len(rows)
        for r in result["validation_rules"]:
            rows.append({
                "object_name": obj, "check_type": "ValidationRule",
                "item_name": r.get("ValidationName") or r.get("Id") or "(unnamed)",
                "is_active": bool(r.get("Active")),
                "direct_hit": bool(r.get("direct_hit")), "detail": r.get("ErrorMessage"),
                "analyzed_at": analyzed_at,
            })
        for r in result["apex_triggers"]:
            usage = ", ".join(
                k.replace("Usage", "") for k in r if k.startswith("Usage") and r.get(k)
            )
            rows.append({
                "object_name": obj, "check_type": "ApexTrigger",
                "item_name": r.get("Name"), "is_active": True, "direct_hit": False,
                "detail": usage, "analyzed_at": analyzed_at,
            })
        for r in result["workflow_rules"]:
            rows.append({
                "object_name": obj, "check_type": "WorkflowRule",
                "item_name": r.get("Name"), "is_active": None, "direct_hit": False,
                "detail": None, "analyzed_at": analyzed_at,
            })
        for r in result["approval_processes"]:
            rows.append({
                "object_name": obj, "check_type": "ApprovalProcess",
                "item_name": r.get("Name"), "is_active": r.get("State") == "Active",
                "direct_hit": False, "detail": r.get("State"), "analyzed_at": analyzed_at,
            })
        for r in result["record_triggered_flows"]:
            rows.append({
                "object_name": obj, "check_type": "RecordTriggeredFlow",
                "item_name": r.get("Label"), "is_active": bool(r.get("IsActive")),
                "direct_hit": False, "detail": r.get("TriggerType"), "analyzed_at": analyzed_at,
            })

        if len(rows) == rows_before:
            # A genuinely clean scan (zero active automation of any kind)
            # would otherwise leave this object with NO rows at all --
            # indistinguishable from "never scanned" for any downstream
            # consumer checking "does this object have automation-risk
            # data on file" (e.g. orchestrator.py's cold-start/tier-3
            # check). One marker row records that the scan itself
            # happened, even though it found nothing.
            rows.append({
                "object_name": obj, "check_type": "ScanCompleted",
                "item_name": "(no active automation found)", "is_active": None,
                "direct_hit": False, "detail": None, "analyzed_at": analyzed_at,
            })

    qualified = sql_dialect.for_engine(engine).qualify(schema, "ObjectAutomationRisk")
    with engine.begin() as cx:
        for result in results:
            cx.execute(
                text(f"DELETE FROM {qualified} WHERE ObjectName = :o"),
                {"o": result["object"]},
            )
        if rows:
            cx.execute(
                text(
                    f"INSERT INTO {qualified} "
                    "(ObjectName, CheckType, ItemName, IsActive, DirectHit, Detail, AnalyzedDate) "
                    "VALUES (:object_name, :check_type, :item_name, :is_active, :direct_hit, :detail, :analyzed_at)"
                ),
                rows,
            )
